#!/usr/bin/env python
"""One-shot importer for the CEO's accrual spreadsheet.

Reads the spreadsheet, creates historical periods seeded from observed rates,
autogenerates uniform cells for active Projects, and overlays per-row overrides
matched by Project.code.
"""

from __future__ import annotations

import argparse
import asyncio
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
from statistics import median

from openpyxl import load_workbook


@dataclass
class SpreadsheetRow:
    type: str
    code: str | None
    pm: str | None
    name: str | None
    value: Decimal
    rate: Decimal
    value_eur: Decimal
    start_date: date | None
    end_date: date | None
    duration: int | None
    monthly: dict[tuple[int, int], Decimal] = field(default_factory=dict)


import re

_SUFFIX_PATTERN = re.compile(r"\.[\w/]+$")


def _normalize_code(code: str | None) -> str | None:
    """Strip ALL whitespace + collapse repeated dots + upper-case.

    Handles divergences such as ``"LSE .TPI2025 .35054413"`` (internal spaces),
    ``"ICIMOD..34229341"`` (doubled dots), and trailing whitespace from copy-
    paste mistakes between Excel and VizzHub.
    """
    if not code:
        return None
    no_ws = "".join(code.split())  # strip all whitespace, internal + external
    collapsed = re.sub(r"\.{2,}", ".", no_ws)
    return collapsed.upper() or None


def _code_prefix(code: str | None) -> str | None:
    """Return the code with its trailing dot-segment stripped, or None.

    Trailing segments are typically Jira issue IDs (``.32232147``), signing
    years (``.24``), or version markers (``.24/25``, ``.2``). Returns None
    when the normalised code has no dot or no trailing alphanumeric segment.
    Only returns a *changed* form (no-op when stripping wouldn't shorten).
    """
    norm = _normalize_code(code)
    if not norm:
        return None
    stripped = _SUFFIX_PATTERN.sub("", norm)
    return stripped if stripped and stripped != norm else None


_AMBIGUOUS = object()  # sentinel for prefix collisions


def _build_excel_index(
    rows: list[SpreadsheetRow],
) -> tuple[dict[str, SpreadsheetRow], dict[str, SpreadsheetRow | object]]:
    """Index Excel rows by normalized code + by unique prefix.

    Returned ``exact_by_code`` keys are full normalized codes. ``prefix_by_code``
    maps prefix → row only when the prefix is unique across rows AND doesn't
    collide with a full normalized code; collisions are marked ``_AMBIGUOUS``.
    """
    exact_by_code: dict[str, SpreadsheetRow] = {}
    for r in rows:
        full = _normalize_code(r.code)
        if full and full not in exact_by_code:
            exact_by_code[full] = r

    prefix_by_code: dict[str, SpreadsheetRow | object] = {}
    for r in rows:
        p = _code_prefix(r.code)
        if not p or p in exact_by_code:
            continue
        if p in prefix_by_code:
            prefix_by_code[p] = _AMBIGUOUS
        else:
            prefix_by_code[p] = r
    return exact_by_code, prefix_by_code


def _db_unique_prefixes(projects: list) -> set[str]:
    """Return prefixes that identify a single DB project."""
    counts: dict[str, int] = defaultdict(int)
    for p in projects:
        prefix = _code_prefix(p.code)
        if prefix:
            counts[prefix] += 1
    return {prefix for prefix, c in counts.items() if c == 1}


def _match_project_to_row(
    project,
    *,
    exact_by_code: dict[str, SpreadsheetRow],
    prefix_by_code: dict[str, SpreadsheetRow | object],
    db_unique_prefixes: set[str],
) -> SpreadsheetRow | None:
    """Resolve a DB project to its Excel row via exact-then-prefix matching.

    Resolution order:
    1. Exact match on normalized code.
    2. DB project's prefix-stripped form matches an Excel full code, provided
       this prefix uniquely identifies one DB project.
    3. DB project's full normalized code is a unique Excel prefix.
    """
    norm = _normalize_code(project.code)
    if norm and norm in exact_by_code:
        return exact_by_code[norm]
    proj_prefix = _code_prefix(project.code)
    if proj_prefix and proj_prefix in db_unique_prefixes and proj_prefix in exact_by_code:
        return exact_by_code[proj_prefix]
    if norm and norm in prefix_by_code:
        candidate = prefix_by_code[norm]
        if candidate is not _AMBIGUOUS:
            return candidate  # type: ignore[return-value]
    return None


def parse_spreadsheet(path: Path) -> list[SpreadsheetRow]:
    """Parse the CEO's accrual workbook into structured rows.

    Sheet layout (per inspection 2026-05-22): sheet 'Sales', year row 5,
    month/header row 6, data rows from row 7. Monthly columns start at 13.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb["Sales"]

    year_by_col: dict[int, int] = {}
    month_by_col: dict[int, int] = {}
    for c in range(13, ws.max_column + 1):
        y = ws.cell(row=5, column=c).value
        m = ws.cell(row=6, column=c).value
        if isinstance(y, (int, float)) and isinstance(m, (int, float)):
            year_by_col[c] = int(y)
            month_by_col[c] = int(m)

    rows: list[SpreadsheetRow] = []
    for r in range(7, ws.max_row + 1):
        type_v = ws.cell(row=r, column=1).value
        if not type_v:
            continue
        rate_v = ws.cell(row=r, column=7).value
        if rate_v is None:
            continue

        monthly: dict[tuple[int, int], Decimal] = {}
        for c, y in year_by_col.items():
            m = month_by_col[c]
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                monthly[(y, m)] = Decimal(str(v)).quantize(Decimal("0.01"))

        def _date(v: object) -> date | None:
            if v is None:
                return None
            return v.date() if hasattr(v, "date") else v  # type: ignore[return-value]

        rows.append(
            SpreadsheetRow(
                type=str(type_v),
                code=str(ws.cell(row=r, column=3).value)
                if ws.cell(row=r, column=3).value
                else None,
                pm=ws.cell(row=r, column=4).value,
                name=ws.cell(row=r, column=5).value,
                value=Decimal(str(ws.cell(row=r, column=6).value or 0)),
                rate=Decimal(str(rate_v)),
                value_eur=Decimal(str(ws.cell(row=r, column=8).value or 0)),
                start_date=_date(ws.cell(row=r, column=9).value),
                end_date=_date(ws.cell(row=r, column=10).value),
                duration=ws.cell(row=r, column=11).value,
                monthly=monthly,
            )
        )
    return rows


async def bootstrap_periods(
    db: AsyncSession,
    rows: list[SpreadsheetRow],
    *,
    current_year: int | None = None,
) -> list:
    """Create one accrual period per year spanned by billable DB projects.

    Year range, currencies and rates are derived from DB state + Excel data
    — no hardcoding. For each (year, non-EUR currency) the period's fx rate
    is the median of ``row.rate`` over spreadsheet rows whose project (matched
    by ``code``) starts that year with that currency. Currencies absent in a
    given year inherit from the previous period via ``create_period``'s merge.
    """
    from sqlalchemy import select
    from sqlalchemy.ext.asyncio import AsyncSession  # noqa: F401

    from app.core.models.project import ProjectDB
    from app.core.services.exchange_rate_service import currency_to_code
    from app.modules.accrual.services import period_service

    proj_result = await db.execute(
        select(ProjectDB).where(
            ProjectDB.is_billable.is_(True),
            ProjectDB.start_date.is_not(None),
            ProjectDB.end_date.is_not(None),
        )
    )
    projects = list(proj_result.scalars().all())
    if not projects:
        return []

    exact_by_code, prefix_by_code = _build_excel_index(rows)
    db_unique_prefixes = _db_unique_prefixes(projects)

    by_year_cur: dict[int, dict[str, list[Decimal]]] = defaultdict(lambda: defaultdict(list))
    for project in projects:
        if not project.currency:
            continue
        currency = currency_to_code(project.currency)
        if not currency or currency == "EUR":
            continue
        row = _match_project_to_row(
            project,
            exact_by_code=exact_by_code,
            prefix_by_code=prefix_by_code,
            db_unique_prefixes=db_unique_prefixes,
        )
        if row is None or row.rate is None or row.start_date is None:
            continue
        by_year_cur[row.start_date.year][currency].append(row.rate)

    if current_year is None:
        current_year = date.today().year
    min_year = min(min(p.start_date.year for p in projects), current_year)
    # Cap year range at current_year: future cells live under the open period
    # until the CEO creates the next period via the UI in due time.
    years = list(range(min_year, current_year + 1))

    from app.modules.accrual.models.accrual_period import AccrualPeriodDB

    created = []
    for y in years:
        fx_rates: dict[str, str] = {}
        for cur, rates in by_year_cur.get(y, {}).items():
            fx_rates[cur] = str(median(rates).quantize(Decimal("0.000001")))

        existing = (
            await db.execute(
                select(AccrualPeriodDB).where(AccrualPeriodDB.start_date == date(y, 1, 1))
            )
        ).scalar_one_or_none()
        if existing is not None:
            created.append(existing)
            continue

        period = await period_service.create_period(
            db,
            start_date=date(y, 1, 1),
            fx_rates_input=fx_rates,
            created_by=None,
        )
        created.append(period)
    return created


async def import_projects(db: AsyncSession, rows: list[SpreadsheetRow]) -> dict:
    """Apply Excel data to DB projects matched by code.

    For each match: set original_budget (if absent), redistribute uniformly
    over the full project range, then overlay per-month overrides where the
    Excel value diverges from the uniform split by more than
    OVERRIDE_THRESHOLD_EUR. Does NOT touch locked_fx_rate — that is a
    CEO-managed transitional override set via the UI.
    """
    from sqlalchemy import select

    from app.core.models.project import ProjectDB
    from app.modules.accrual.constants import OVERRIDE_THRESHOLD_EUR
    from app.modules.accrual.models.project_accrual_cell import ProjectAccrualCellDB
    from app.modules.accrual.services import cell_service

    report: dict = {
        "matched": 0,
        "original_budget_set": 0,
        "overrides_imported": 0,
        "unmatched": [],
    }
    exact_by_code, prefix_by_code = _build_excel_index(rows)

    proj_result = await db.execute(
        select(ProjectDB).where(ProjectDB.status.in_(["proposal", "live", "finished"]))
    )
    projects = list(proj_result.scalars().all())
    db_unique_prefixes = _db_unique_prefixes(projects)
    matched_rows: set[int] = set()  # track id() of matched SpreadsheetRow objects

    for project in projects:
        if not project.start_date or not project.end_date:
            continue
        row = _match_project_to_row(
            project,
            exact_by_code=exact_by_code,
            prefix_by_code=prefix_by_code,
            db_unique_prefixes=db_unique_prefixes,
        )
        if row is None:
            continue
        report["matched"] += 1
        matched_rows.add(id(row))

        if project.original_budget is None and row.value:
            project.original_budget = row.value
            report["original_budget_set"] += 1

        await db.flush()

        # full_range=True: populate cells across the entire project lifespan,
        # bypassing the active-period clip (current period may be 2026+).
        await cell_service.redistribute_for_project(db, project_id=project.id, full_range=True)

        existing_result = await db.execute(
            select(ProjectAccrualCellDB).where(ProjectAccrualCellDB.project_id == project.id)
        )
        existing = {(c.year, c.month): c for c in existing_result.scalars().all()}
        for (y, m), spreadsheet_amount in row.monthly.items():
            cell = existing.get((y, m))
            if cell is None or cell.is_frozen:
                continue
            uniform = cell.amount
            if abs(spreadsheet_amount - uniform) > OVERRIDE_THRESHOLD_EUR:
                await cell_service.set_cell_amount(
                    db,
                    project_id=project.id,
                    year=y,
                    month=m,
                    amount=spreadsheet_amount,
                )
                report["overrides_imported"] += 1

    for r in rows:
        if r.code and id(r) not in matched_rows:
            report["unmatched"].append({"code": r.code, "name": r.name, "type": r.type})

    return report


async def freeze_historical_periods(db: AsyncSession) -> int:
    """Re-run the freeze pass on every closed period.

    Periods are created in bootstrap_periods before any cells exist, so the
    initial close freezes nothing. After import_projects populates cells,
    this step retroactively freezes them with each period's resolved rate.
    """
    from sqlalchemy import select

    from app.modules.accrual.models.accrual_period import AccrualPeriodDB
    from app.modules.accrual.services import period_service

    result = await db.execute(
        select(AccrualPeriodDB)
        .where(AccrualPeriodDB.status == "closed")
        .order_by(AccrualPeriodDB.start_date)
    )
    total = 0
    for period in result.scalars().all():
        total += await period_service.freeze_period_cells(db, period_id=period.id)
    return total


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--spreadsheet", required=True, type=Path)
    p.add_argument("--dry-run", action="store_true")
    p.add_argument(
        "--periods-only",
        action="store_true",
        help="Stop after creating historical periods (review rates before full import).",
    )
    return p.parse_args(argv)


class _DryRun(Exception):
    """Sentinel to trigger db.begin() rollback for --dry-run."""


async def main_async(args: argparse.Namespace) -> int:
    if not args.spreadsheet.exists():
        print(f"[importer] file not found: {args.spreadsheet}", file=sys.stderr)
        return 1

    from app.database import async_session_maker

    rows = parse_spreadsheet(args.spreadsheet)
    print(f"[importer] parsed {len(rows)} rows from {args.spreadsheet}")

    async with async_session_maker() as db:
        try:
            async with db.begin():
                periods = await bootstrap_periods(db, rows)
                print(f"[importer] created {len(periods)} periods")
                if args.periods_only:
                    if args.dry_run:
                        raise _DryRun()
                    return 0
                report = await import_projects(db, rows)
                frozen = await freeze_historical_periods(db)
                print(f"[importer] {report} | frozen_cells={frozen}")
                if args.dry_run:
                    raise _DryRun()
        except _DryRun:
            print("[importer] dry-run complete — rolled back", file=sys.stderr)
    return 0


def main() -> int:
    return asyncio.run(main_async(parse_args()))


if __name__ == "__main__":
    sys.exit(main())
