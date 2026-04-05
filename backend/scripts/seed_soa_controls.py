"""Seed the Statement of Applicability registry with all 93 ISO 27001:2022 Annex A controls.

Reads existing data from the SOA Excel file and maps each control to its official
English name (per ISO 27001:2022 / ISO 27002:2022).

Run: python -m scripts.seed_soa_controls
Idempotent: skips if rows already exist for the SOA node.
"""

from __future__ import annotations

import asyncio
import re
from pathlib import Path

import structlog
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.config import get_settings
from app.database import Base  # noqa: F401
from app.core.models.user import UserDB  # noqa: F401
from app.modules.iso_docs.models.node import IsoDocNodeDB
from app.modules.iso_docs.models.registry_row import RegistryRowDB
from app.modules.iso_docs.models.registry_type import RegistryTypeDB

logger = structlog.get_logger()

SOA_SLUG = "statement-of-applicability"

# Official ISO 27001:2022 Annex A control names (English)
# Sources: ISO 27001:2022 standard, cross-verified with iseoblue.com and dataguard.com
ANNEX_A_CONTROLS_EN: dict[str, str] = {
    # A.5 — Organizational controls (37)
    "5.1": "Policies for information security",
    "5.2": "Information security roles and responsibilities",
    "5.3": "Segregation of duties",
    "5.4": "Management responsibilities",
    "5.5": "Contact with authorities",
    "5.6": "Contact with special interest groups",
    "5.7": "Threat intelligence",
    "5.8": "Information security in project management",
    "5.9": "Inventory of information and other associated assets",
    "5.10": "Acceptable use of information and other associated assets",
    "5.11": "Return of assets",
    "5.12": "Classification of information",
    "5.13": "Labelling of information",
    "5.14": "Information transfer",
    "5.15": "Access control",
    "5.16": "Identity management",
    "5.17": "Authentication information",
    "5.18": "Access rights",
    "5.19": "Information security in supplier relationships",
    "5.20": "Addressing information security within supplier agreements",
    "5.21": "Managing information security in the ICT supply chain",
    "5.22": "Monitoring, review and change management of supplier services",
    "5.23": "Information security for use of cloud services",
    "5.24": "Information security incident management planning and preparation",
    "5.25": "Assessment and decision on information security events",
    "5.26": "Response to information security incidents",
    "5.27": "Learning from information security incidents",
    "5.28": "Collection of evidence",
    "5.29": "Information security during disruption",
    "5.30": "ICT readiness for business continuity",
    "5.31": "Legal, statutory, regulatory and contractual requirements",
    "5.32": "Intellectual property rights",
    "5.33": "Protection of records",
    "5.34": "Privacy and protection of PII",
    "5.35": "Independent review of information security",
    "5.36": "Compliance with policies, rules and standards for information security",
    "5.37": "Documented operating procedures",
    # A.6 — People controls (8)
    "6.1": "Screening",
    "6.2": "Terms and conditions of employment",
    "6.3": "Information security awareness, education and training",
    "6.4": "Disciplinary process",
    "6.5": "Responsibilities after termination or change of employment",
    "6.6": "Confidentiality or non-disclosure agreements",
    "6.7": "Remote working",
    "6.8": "Information security event reporting",
    # A.7 — Physical controls (14)
    "7.1": "Physical security perimeters",
    "7.2": "Physical entry",
    "7.3": "Securing offices, rooms and facilities",
    "7.4": "Physical security monitoring",
    "7.5": "Protecting against physical and environmental threats",
    "7.6": "Working in secure areas",
    "7.7": "Clear desk and clear screen",
    "7.8": "Equipment siting and protection",
    "7.9": "Security of assets off-premises",
    "7.10": "Storage media",
    "7.11": "Supporting utilities",
    "7.12": "Cabling security",
    "7.13": "Equipment maintenance",
    "7.14": "Secure disposal or reuse of equipment",
    # A.8 — Technological controls (34)
    "8.1": "User endpoint devices",
    "8.2": "Privileged access rights",
    "8.3": "Information access restriction",
    "8.4": "Access to source code",
    "8.5": "Secure authentication",
    "8.6": "Capacity management",
    "8.7": "Protection against malware",
    "8.8": "Management of technical vulnerabilities",
    "8.9": "Configuration management",
    "8.10": "Information deletion",
    "8.11": "Data masking",
    "8.12": "Data leakage prevention",
    "8.13": "Information backup",
    "8.14": "Redundancy of information processing facilities",
    "8.15": "Logging",
    "8.16": "Monitoring activities",
    "8.17": "Clock synchronization",
    "8.18": "Use of privileged utility programs",
    "8.19": "Installation of software on operational systems",
    "8.20": "Networks security",
    "8.21": "Security of network services",
    "8.22": "Segregation of networks",
    "8.23": "Web filtering",
    "8.24": "Use of cryptography",
    "8.25": "Secure development life cycle",
    "8.26": "Application security requirements",
    "8.27": "Secure system architecture and engineering principles",
    "8.28": "Secure coding",
    "8.29": "Security testing in development and acceptance",
    "8.30": "Outsourced development",
    "8.31": "Separation of development, test and production environments",
    "8.32": "Change management",
    "8.33": "Test information",
    "8.34": "Protection of information systems during audit testing",
}

CATEGORY_MAP: dict[str, str] = {
    "5": "Organizational",
    "6": "People",
    "7": "Physical",
    "8": "Technological",
}

CONTROL_TYPE_MAP = {
    "preventivo": "Preventive",
    "detectivo": "Detective",
    "correctivo": "Corrective",
}


def _extract_code(raw: str) -> str:
    """Extract the numeric control code from Excel text like '5.1 Políticas de...'."""
    match = re.match(r"(\d+\.\d+)", raw.strip())
    return match.group(1) if match else raw.strip()


def _extract_name_es(raw: str) -> str:
    """Extract just the Spanish name, stripping the leading code."""
    match = re.match(r"\d+\.\d+\s+(.*)", raw.strip())
    return match.group(1).strip() if match else raw.strip()


def _parse_control_type(raw: str) -> str | None:
    """Parse control type, picking the first recognized type."""
    lower = raw.lower().strip()
    for es_term, en_term in CONTROL_TYPE_MAP.items():
        if es_term in lower:
            return en_term
    return None


def _parse_applicable(raw: str) -> bool:
    return "aplica" in raw.lower() and "exclu" not in raw.lower()


def _clean_text(val: str) -> str | None:
    """Strip and return None for empty strings."""
    cleaned = val.strip()
    return cleaned if cleaned else None


def load_excel_data() -> list[dict]:
    """Load SOA controls from the Excel file."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed, using hardcoded controls only")
        return []

    excel_path = Path(__file__).parent.parent.parent / (
        "temp/iso/REGISTROS/ANEXO A - DECLARACIÓN APLICABILIDAD/"
        "SOA Declaracion de aplicabilidad.xlsx"
    )
    if not excel_path.exists():
        logger.warning("soa_excel_not_found", path=str(excel_path))
        return []

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb["Controles"]

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        code_cell = row[0].value
        if code_cell is None or str(code_cell).strip() == "Código":
            continue

        vals = [str(c.value).strip() if c.value is not None else "" for c in row[:12]]
        code = _extract_code(vals[0])

        rows.append({
            "code": code,
            "name_es": _extract_name_es(vals[0]),
            "description_es": vals[1],
            "control_type": _parse_control_type(vals[2]),
            "applicable": _parse_applicable(vals[7]),
            "justification": _clean_text(vals[8]),
            "management": _clean_text(vals[9]),
            "evidence": _clean_text(vals[10]),
            "notes": _clean_text(vals[11]),
        })

    return rows


def build_soa_rows() -> list[dict]:
    """Build the complete list of 93 SOA row data dicts."""
    excel_data = load_excel_data()
    excel_by_code = {r["code"]: r for r in excel_data}

    rows: list[dict] = []
    for code, name_en in ANNEX_A_CONTROLS_EN.items():
        control_id = f"A.{code}"
        excel_row = excel_by_code.get(code, {})

        management_text = excel_row.get("management")
        has_management = bool(management_text)
        impl_status = "Implemented" if has_management else "Planned"
        if not excel_row.get("applicable", True):
            impl_status = "N/A"

        section = code.split(".")[0]
        rows.append({
            "control_id": control_id,
            "category": CATEGORY_MAP[section],
            "control_name": name_en,
            "control_name_es": excel_row.get("name_es", ""),
            "control_type": excel_row.get("control_type"),
            "applicable": excel_row.get("applicable", True),
            "justification": excel_row.get("justification"),
            "implementation_status": impl_status,
            "management": management_text,
            "evidence": excel_row.get("evidence"),
            "notes": excel_row.get("notes"),
        })

    return rows


async def seed(db_url: str | None = None) -> int:
    url = db_url or get_settings().database_url
    engine = create_async_engine(url)
    session_maker = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with session_maker() as db:
        # 1. Find the SOA registry type
        result = await db.execute(
            select(RegistryTypeDB).where(RegistryTypeDB.slug == SOA_SLUG)
        )
        soa_type = result.scalar_one_or_none()
        if not soa_type:
            print(f"  ERROR: Registry type '{SOA_SLUG}' not found. Run seed_registry_types first.")
            await engine.dispose()
            return 0

        # Update schema + description if needed
        from scripts.seed_registry_types import REGISTRY_TYPES

        soa_def = next(r for r in REGISTRY_TYPES if r["slug"] == SOA_SLUG)
        soa_type.schema = soa_def["schema"]
        soa_type.description = soa_def["description"]
        soa_type.default_sort_key = soa_def.get("default_sort_key")

        # 2. Find or create the SOA node
        result = await db.execute(
            select(IsoDocNodeDB).where(
                IsoDocNodeDB.registry_type_id == soa_type.id,
                IsoDocNodeDB.type == "registry",
            )
        )
        soa_node = result.scalar_one_or_none()

        if not soa_node:
            # Find the "Registries" group to place it under
            result = await db.execute(
                select(IsoDocNodeDB).where(
                    IsoDocNodeDB.type == "group",
                    IsoDocNodeDB.slug == "registries",
                )
            )
            registries_group = result.scalar_one_or_none()

            max_pos = (await db.execute(
                select(func.coalesce(func.max(IsoDocNodeDB.position), -1)).where(
                    IsoDocNodeDB.parent_id == (registries_group.id if registries_group else None)
                )
            )).scalar() or 0

            soa_node = IsoDocNodeDB(
                title="Statement of Applicability",
                slug="statement-of-applicability",
                type="registry",
                parent_id=registries_group.id if registries_group else None,
                position=max_pos + 1,
                registry_type_id=soa_type.id,
            )
            db.add(soa_node)
            await db.flush()
            print(f"  created node: {soa_node.title} (id={soa_node.id})")
        else:
            print(f"  existing node: {soa_node.title} (id={soa_node.id})")

        # 3. Check if rows already exist
        row_count = (await db.execute(
            select(func.count()).where(RegistryRowDB.node_id == soa_node.id)
        )).scalar() or 0

        created = 0
        if row_count > 0:
            print(f"  skip: {row_count} rows already exist for SOA node")
        else:
            # 4. Seed the 93 controls
            soa_rows = build_soa_rows()
            for idx, row_data in enumerate(soa_rows):
                db.add(RegistryRowDB(
                    node_id=soa_node.id,
                    row_index=idx,
                    data=row_data,
                ))
            created = len(soa_rows)
            print(f"  seeded {created} Annex A controls")

        await db.commit()

    await engine.dispose()
    return created


async def main() -> None:
    print("Seeding SOA controls...")
    count = await seed()
    if count:
        print(f"Done. Seeded {count} controls.")
    else:
        print("Done. No new rows created.")


if __name__ == "__main__":
    asyncio.run(main())
