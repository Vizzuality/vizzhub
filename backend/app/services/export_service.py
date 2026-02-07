"""XLSX export service for project scorecard data."""

from collections.abc import Callable
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select, tuple_
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import ScoringConfig
from app.models.global_metrics import GlobalMetricsDB, GlobalMetricsRecord
from app.models.metrics import MetricsCreate, MetricsDB
from app.models.project import ProjectDB
from app.services.export_definitions import DIMENSION_DEFINITIONS, get_metric_rows
from app.services.export_helpers import (
    DEFAULT_GREEN_THRESHOLD,
    DEFAULT_YELLOW_THRESHOLD,
    THIN_BORDER,
    apply_header_style,
    apply_indicator_traffic_light,
    apply_row_style,
    apply_score_traffic_light,
    create_methodology_sheet,
    format_month_header,
    freeze_panes,
    set_column_widths,
)
from app.services.score_computation import ScoreComputationService


class ExportService:
    """Generates XLSX exports for scorecard data."""

    def __init__(self, config: ScoringConfig):
        self.config = config
        self.score_service = ScoreComputationService(config)
        self._green = self._get_threshold("threshold_green", DEFAULT_GREEN_THRESHOLD)
        self._yellow = self._get_threshold("threshold_yellow", DEFAULT_YELLOW_THRESHOLD)

    async def export_project_detail(
        self,
        db: AsyncSession,
        project: ProjectDB,
        start_year: int,
        start_month: int,
        end_year: int,
        end_month: int,
        snapshot_type: str = "cumulative",
    ) -> BytesIO:
        """Generate project detail XLSX with Scorecard and Methodology sheets."""
        periods = self._generate_periods(start_year, start_month, end_year, end_month)
        metrics_by_period = await self._get_metrics_by_period(
            db, project.id, periods, snapshot_type
        )
        scores_by_period = self._compute_scores(metrics_by_period)

        wb = Workbook()
        ws = wb.active
        ws.title = "Scorecard"

        self._write_project_summary(ws, project, snapshot_type)
        ws.append([])
        self._write_metrics_table(
            ws, periods, lambda key, level, period: self._extract_value(
                key, level, scores_by_period.get(period)
            )
        )
        self._apply_scorecard_widths(ws, periods)

        create_methodology_sheet(wb, self.config)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return self._save_to_bytes(wb)

    async def export_global_dashboard(
        self,
        db: AsyncSession,
        start_year: int,
        start_month: int,
        end_year: int,
        end_month: int,
        snapshot_type: str = "cumulative",
    ) -> BytesIO:
        """Generate global dashboard XLSX with Scorecard and Methodology sheets.

        Uses pre-computed GlobalMetricsDB records (averaged across all projects).
        """
        periods = self._generate_periods(start_year, start_month, end_year, end_month)
        global_by_period = await self._get_global_metrics_by_period(db, periods)

        wb = Workbook()
        ws = wb.active
        ws.title = "Scorecard"

        self._write_global_summary(ws, periods, global_by_period)
        ws.append([])
        self._write_metrics_table(
            ws, periods, lambda key, level, period: self._extract_global_value(
                key, level, global_by_period.get(period)
            )
        )
        self._apply_scorecard_widths(ws, periods)

        create_methodology_sheet(wb, self.config)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return self._save_to_bytes(wb)

    # --- Data fetching ---

    async def _get_metrics_by_period(
        self,
        db: AsyncSession,
        project_id: UUID,
        periods: list[tuple[int, int]],
        snapshot_type: str,
    ) -> dict[tuple[int, int], MetricsDB | None]:
        """Fetch metrics for each period, returns dict of (year, month) -> MetricsDB."""
        result: dict[tuple[int, int], MetricsDB | None] = dict.fromkeys(periods)
        if not periods:
            return result

        query = (
            select(MetricsDB)
            .where(MetricsDB.project_id == project_id)
            .where(MetricsDB.snapshot_type == snapshot_type)
            .where(
                tuple_(MetricsDB.period_year, MetricsDB.period_month).in_(periods)
            )
            .order_by(MetricsDB.created_at.desc())
        )
        res = await db.execute(query)
        for row in res.scalars().all():
            key = (row.period_year, row.period_month)
            if result[key] is None:
                result[key] = row
        return result

    async def _get_global_metrics_by_period(
        self,
        db: AsyncSession,
        periods: list[tuple[int, int]],
    ) -> dict[tuple[int, int], GlobalMetricsRecord | None]:
        """Fetch pre-computed global metrics for each period."""
        result: dict[tuple[int, int], GlobalMetricsRecord | None] = dict.fromkeys(periods)
        if not periods:
            return result

        query = (
            select(GlobalMetricsDB)
            .where(
                tuple_(
                    GlobalMetricsDB.period_year, GlobalMetricsDB.period_month
                ).in_(periods)
            )
        )
        res = await db.execute(query)
        for row in res.scalars().all():
            key = (row.period_year, row.period_month)
            if key in result:
                result[key] = GlobalMetricsRecord.from_db(row)
        return result

    # --- Score computation ---

    def _compute_scores(
        self,
        metrics_by_period: dict[tuple[int, int], MetricsDB | None],
    ) -> dict[tuple[int, int], dict | None]:
        """Compute scores for each period that has metrics."""
        result: dict[tuple[int, int], dict | None] = {}
        for period, metrics_db in metrics_by_period.items():
            if metrics_db is None:
                result[period] = None
                continue
            metrics = MetricsCreate.from_db(metrics_db)
            indicators, scores = self.score_service.compute(
                metrics, sev1_incident=metrics_db.sev1_incident
            )
            result[period] = {
                "indicators": indicators,
                "scores": scores,
            }
        return result

    # --- Sheet builders ---

    @staticmethod
    def _write_project_summary(
        ws,
        project: ProjectDB,
        snapshot_type: str,
    ) -> None:
        """Write project info rows at the top of the active sheet."""
        info = [
            ("Project", project.name),
            ("Jira Key", project.jira_project_key or "-"),
            ("GitHub Repo", project.github_repo or "-"),
            ("Status", project.status),
            ("Start Date", str(project.start_date) if project.start_date else "-"),
            ("End Date", str(project.end_date) if project.end_date else "-"),
            ("Snapshot Type", snapshot_type),
        ]
        for label, value in info:
            ws.append([label, value])

    def _write_global_summary(
        self,
        ws,
        periods: list[tuple[int, int]],
        global_by_period: dict[tuple[int, int], GlobalMetricsRecord | None],
    ) -> None:
        """Write global summary rows at the top of the active sheet."""
        ws.append(["Global Dashboard", "Averaged scores across all projects"])
        ws.append([])

        header = [""] + [format_month_header(y, m) for y, m in periods]
        ws.append(header)
        apply_header_style(ws, ws.max_row)

        counts = ["Projects"]
        for period in periods:
            record = global_by_period.get(period)
            counts.append(record.project_count if record else None)
        ws.append(counts)

        scores = ["Overall Score"]
        for period in periods:
            record = global_by_period.get(period)
            scores.append(
                round(record.scores.score.value, 1)
                if record and record.scores.score.value is not None
                else None
            )
        ws.append(scores)

        score_row = ws.max_row
        for col_idx in range(2, 2 + len(periods)):
            cell = ws.cell(row=score_row, column=col_idx)
            cell.border = THIN_BORDER
            apply_score_traffic_light(
                cell, cell.value, self._green, self._yellow
            )

        for dim_def in DIMENSION_DEFINITIONS:
            dim_key = dim_def["key"]
            row = [dim_def["name"]]
            for period in periods:
                record = global_by_period.get(period)
                value = None
                if record:
                    score_val = getattr(record.scores, dim_key, None)
                    if score_val and score_val.value is not None:
                        value = round(score_val.value, 1)
                row.append(value)
            ws.append(row)

            current_row = ws.max_row
            for col_idx in range(2, 2 + len(periods)):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = THIN_BORDER
                apply_score_traffic_light(
                    cell, cell.value, self._green, self._yellow
                )

    def _write_metrics_table(
        self,
        ws,
        periods: list[tuple[int, int]],
        extract_fn: Callable[[str, int, tuple[int, int]], float | int | None],
    ) -> None:
        """Write hierarchical metrics table with traffic-light coloring onto ws.

        Appends rows to the current sheet starting at the next available row.
        """
        metric_rows = get_metric_rows()

        header = ["Name", "Description", "Formula", "Target"]
        for year, month in periods:
            header.append(format_month_header(year, month))
        ws.append(header)
        header_row = ws.max_row
        apply_header_style(ws, header_row)

        for metric_row in metric_rows:
            level = metric_row["level"]
            indent = "  " * level
            target = self._get_target_for_metric(metric_row["key"], level)

            row_data = [
                f"{indent}{metric_row['name']}",
                metric_row["description"],
                metric_row["formula"],
                target if target else "-",
            ]

            for period in periods:
                row_data.append(extract_fn(metric_row["key"], level, period))

            ws.append(row_data)
            current_row = ws.max_row
            apply_row_style(ws, current_row, level)

            for col_idx, _period in enumerate(periods, start=5):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = THIN_BORDER
                if level <= 1:
                    apply_score_traffic_light(
                        cell, cell.value, self._green, self._yellow
                    )
                else:
                    apply_indicator_traffic_light(
                        cell, cell.value,
                        self._green / 100, self._yellow / 100,
                    )

        freeze_panes(ws, header_row + 1, 5)

    @staticmethod
    def _apply_scorecard_widths(ws, periods: list[tuple[int, int]]) -> None:
        """Set column widths for the combined Scorecard sheet."""
        widths = {"A": 35, "B": 50, "C": 45, "D": 12}
        for i, _ in enumerate(periods):
            widths[get_column_letter(5 + i)] = 12
        set_column_widths(ws, widths)

    # --- Helpers ---

    @staticmethod
    def _generate_periods(
        start_year: int, start_month: int, end_year: int, end_month: int
    ) -> list[tuple[int, int]]:
        """Generate a list of (year, month) tuples for the given range."""
        periods: list[tuple[int, int]] = []
        year, month = start_year, start_month
        while (year, month) <= (end_year, end_month):
            periods.append((year, month))
            month += 1
            if month > 12:
                month = 1
                year += 1
        return periods

    def _get_target_for_metric(self, key: str, level: int) -> str | None:
        """Get a display-friendly target string for a metric row."""
        if level == 0:
            return str(int(self._green))
        if level == 1:
            return str(int(self._green))
        try:
            return str(self.config.get_target(key))
        except (KeyError, ValueError):
            return None

    def _get_threshold(self, name: str, default: float) -> float:
        """Get a threshold constant from config, falling back to default."""
        try:
            val = self.config.get_constant(name)
            return val if val > 0 else default
        except (KeyError, ValueError):
            return default

    @staticmethod
    def _extract_value(
        key: str, level: int, score_data: dict | None
    ) -> int | float | None:
        """Extract the appropriate value from computed score data."""
        if score_data is None:
            return None
        if level == 0:
            val = score_data["scores"].score
            return round(val, 1) if val is not None else None
        if level == 1:
            val = getattr(score_data["scores"].dimensions, key, None)
            return round(val, 1) if val is not None else None
        val = getattr(score_data["indicators"], key, None)
        return round(val, 1) if val is not None else None

    @staticmethod
    def _extract_global_value(
        key: str, level: int, record: GlobalMetricsRecord | None
    ) -> float | None:
        """Extract averaged value from a GlobalMetricsRecord."""
        if record is None:
            return None
        if level == 0:
            val = record.scores.score.value
            return round(val, 1) if val is not None else None
        if level == 1:
            score_val = getattr(record.scores, key, None)
            if score_val and score_val.value is not None:
                return round(score_val.value, 1)
            return None
        indicator_val = getattr(record.indicators, key, None)
        if indicator_val and indicator_val.value is not None:
            return round(indicator_val.value, 1)
        return None

    @staticmethod
    def _save_to_bytes(wb: Workbook) -> BytesIO:
        """Save workbook to an in-memory BytesIO buffer."""
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
