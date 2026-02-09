"""Shared XLSX formatting helpers for export service."""

import calendar

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.config import ScoringConfig
from app.services.export_definitions import DIMENSION_DEFINITIONS, INDICATOR_DEFINITIONS

# Soft palette — same hues (#4CAF50, #FFC107, #F44336) at two opacity levels on white.
# Strong (~35% opacity): scores and dimensions (level 0-1)
GREEN_FILL = PatternFill(start_color="C1E3C2", end_color="C1E3C2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFE9A8", end_color="FFE9A8", fill_type="solid")
RED_FILL = PatternFill(start_color="FBBDB9", end_color="FBBDB9", fill_type="solid")

# Subtle (~15% opacity): indicators (level 2)
GREEN_FILL_SUBTLE = PatternFill(start_color="E4F3E5", end_color="E4F3E5", fill_type="solid")
YELLOW_FILL_SUBTLE = PatternFill(start_color="FFF6DA", end_color="FFF6DA", fill_type="solid")
RED_FILL_SUBTLE = PatternFill(start_color="FDE3E1", end_color="FDE3E1", fill_type="solid")

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DIM_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
DIM_FONT = Font(bold=True, size=11)

SCORE_FONT = Font(bold=True, size=12)
SCORE_FILL = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

DEFAULT_GREEN_THRESHOLD = 80
DEFAULT_YELLOW_THRESHOLD = 60


def apply_score_traffic_light(
    cell,
    value: float | int | None,
    green: float = DEFAULT_GREEN_THRESHOLD,
    yellow: float = DEFAULT_YELLOW_THRESHOLD,
) -> None:
    """Apply strong green/yellow/red fill for scores and dimensions (0-100 scale)."""
    if value is None:
        return
    if value >= green:
        cell.fill = GREEN_FILL
    elif value >= yellow:
        cell.fill = YELLOW_FILL
    else:
        cell.fill = RED_FILL


def apply_indicator_traffic_light(
    cell,
    value: float | int | None,
    green: float = 0.80,
    yellow: float = 0.60,
) -> None:
    """Apply subtle green/yellow/red fill for indicators (0-1 scale)."""
    if value is None:
        return
    if value >= green:
        cell.fill = GREEN_FILL_SUBTLE
    elif value >= yellow:
        cell.fill = YELLOW_FILL_SUBTLE
    else:
        cell.fill = RED_FILL_SUBTLE


def apply_header_style(ws: Worksheet, row: int = 1) -> None:
    """Apply dark header style to a row."""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_row_style(ws: Worksheet, row: int, level: int) -> None:
    """Apply style based on hierarchy level."""
    for cell in ws[row]:
        cell.border = THIN_BORDER
        if level == 0:
            cell.font = SCORE_FONT
            cell.fill = SCORE_FILL
        elif level == 1:
            cell.font = DIM_FONT
            cell.fill = DIM_FILL


def format_month_header(year: int, month: int) -> str:
    """Format year/month as 'Jan 2025'."""
    return f"{calendar.month_abbr[month]} {year}"


def set_column_widths(ws: Worksheet, widths: dict[str, int]) -> None:
    """Set column widths from a dict of column letter -> width."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def freeze_panes(ws: Worksheet, row: int, col: int) -> None:
    """Freeze rows above and columns to the left of the given cell."""
    ws.freeze_panes = f"{get_column_letter(col)}{row}"


def create_methodology_sheet(wb: Workbook, config: ScoringConfig) -> Worksheet:
    """Create the Methodology sheet with scoring model explanation and current config."""
    ws = wb.create_sheet("Methodology")

    green_th = _get_threshold(config, "threshold_green", DEFAULT_GREEN_THRESHOLD)
    yellow_th = _get_threshold(config, "threshold_yellow", DEFAULT_YELLOW_THRESHOLD)

    ws.append(["Scoring Methodology"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Scoring Model"])
    ws.cell(row=3, column=1).font = Font(bold=True, size=12)
    ws.append(["Raw Metrics", "->", "Normalized Indicators (0-1)", "->", "Weighted Scores (0-100)"])
    ws.append(["Collectors fetch data from Jira and GitHub. Manual inputs supplement automated data."])
    ws.append(["Indicators are normalized to a 0-1 scale. Scores are weighted averages scaled to 0-100."])
    ws.append([])

    ws.append(["Traffic Light Legend"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    ind_green = green_th / 100
    ind_yellow = yellow_th / 100

    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="Scores & Dimensions (0-100)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)
    ws.cell(row=row, column=2, value="Indicators (0-1)")
    ws.cell(row=row, column=2).font = Font(bold=True, size=10)
    ws.cell(row=row, column=3, value="Category")
    ws.cell(row=row, column=3).font = Font(bold=True, size=10)
    row += 1
    ws.cell(row=row, column=1, value=f"Score >= {green_th:.0f}").fill = GREEN_FILL
    ws.cell(row=row, column=2, value=f"Value >= {ind_green:.2f}").fill = GREEN_FILL_SUBTLE
    ws.cell(row=row, column=3, value="Good performance")
    row += 1
    ws.cell(row=row, column=1, value=f"Score >= {yellow_th:.0f}").fill = YELLOW_FILL
    ws.cell(row=row, column=2, value=f"Value >= {ind_yellow:.2f}").fill = YELLOW_FILL_SUBTLE
    ws.cell(row=row, column=3, value="At risk / needs attention")
    row += 1
    ws.cell(row=row, column=1, value=f"Score < {yellow_th:.0f}").fill = RED_FILL
    ws.cell(row=row, column=2, value=f"Value < {ind_yellow:.2f}").fill = RED_FILL_SUBTLE
    ws.cell(row=row, column=3, value="Critical / poor performance")
    ws.append([])

    ws.append(["Snapshot Types"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Cumulative", "Project-to-date metrics from start_date to period end"])
    ws.append(["Punctual", "Single month metrics for that period only"])
    ws.append([])

    ws.append(["Global Dimension Weights"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Dimension", "Weight"])
    dims = ["time", "cost", "quality", "value", "satisfaction", "flow", "engineering", "risk"]
    for dim in dims:
        weight = config.get_weight("global", dim)
        ws.append([f"P_{dim}", f"{weight:.0%}"])
    ws.append([])

    ws.append(["KPI Reference"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    headers = ["Dimension", "Indicator", "Description", "Formula", "Target", "Weight"]
    ws.append(headers)
    header_row = ws.max_row
    apply_header_style(ws, header_row)

    for dim_def in DIMENSION_DEFINITIONS:
        dim_key = dim_def["key"].replace("p_", "")
        for ind_key in dim_def["indicators"]:
            ind = INDICATOR_DEFINITIONS[ind_key]
            target = _safe_get_target(config, ind_key)
            weight = _safe_get_weight(config, dim_key, ind_key)
            ws.append([
                dim_def["name"],
                ind["name"],
                ind["description"],
                ind["formula"],
                target,
                weight,
            ])

    set_column_widths(ws, {"A": 25, "B": 28, "C": 50, "D": 60, "E": 12, "F": 12})

    return ws


def _get_threshold(config: ScoringConfig, name: str, default: float) -> float:
    """Get a threshold constant from config, falling back to default."""
    try:
        val = config.get_constant(name)
        return val if val > 0 else default
    except (KeyError, ValueError):
        return default


def _safe_get_target(config: ScoringConfig, indicator_key: str) -> str:
    """Get target value as string, return '-' if not configured."""
    try:
        val = config.get_target(indicator_key)
        if val is not None and abs(val) < 1e-9:
            return "-"
        return str(val)
    except (KeyError, ValueError):
        return "-"


def _safe_get_weight(config: ScoringConfig, _dim_key: str, ind_key: str) -> str:
    """Get weight value as string, return '-' if not configured."""
    weight_map = {
        "spi": ("time", "spi"),
        "on_time_milestones": ("time", "milestones"),
        "cpi": ("cost", "cpi"),
        "budget_variance": ("cost", "variance"),
        "defect_density": ("quality", "defect_density"),
        "escaped_rate": ("quality", "escaped_rate"),
        "mttr_hours": ("quality", "mttr"),
        "governance_compliance": ("quality", "governance"),
        "story_review_ratio": ("quality", "story_review"),
        "pr_review_ratio": ("quality", "pr_review"),
        "change_failure_rate": ("quality", "change_failure_rate"),
        "post_contract_tasks": ("quality", "post_contract_tasks"),
        "okr_impact": ("value", "okr_impact"),
        "pm_satisfaction": ("satisfaction", "pm_estimation"),
        "client_satisfaction": ("satisfaction", "client_survey"),
        "lead_time_days": ("flow", "lead_time"),
        "commitment_reliability": ("flow", "commitment_reliability"),
        "pr_size_median": ("flow", "pr_size"),
        "review_turnaround_hours": ("flow", "review_turnaround"),
        "deployment_frequency": ("flow", "deployment_frequency"),
        "test_maturity": ("engineering", "test_maturity"),
        "arch_checklist": ("engineering", "architecture"),
        "prs_without_review": ("risk", "pr_no_review"),
        "high_vulns": ("risk", "high_vulns"),
    }
    mapping = weight_map.get(ind_key)
    if not mapping:
        return "-"
    try:
        val = config.get_weight(mapping[0], mapping[1])
        return f"{val:.0%}"
    except (KeyError, ValueError):
        return "-"
