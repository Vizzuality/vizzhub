"""Portfolio Overview import: parse xlsx -> staging, match, apply decisions."""

import io
from dataclasses import dataclass
from datetime import UTC, datetime
from uuid import UUID

import structlog
from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.sql import func

from app.core.api.clients import slugify
from app.core.models.client import ClientDB
from app.core.models.portfolio_overview import (
    PortfolioOverviewStagingDB,
    PortfolioProfileDB,
    ProgramAction,
)
from app.core.models.program import ProgramDB
from app.core.models.project import ProjectDB
from app.core.models.taxonomy import EntityTermDB, TaxonomyDB, TaxonomyTermDB
from app.core.services.name_matching import rank

logger = structlog.get_logger()

_SHEET = "Categorised"
_HEADER_ROW = 2
_SEPARATOR = "only old projects"

# 1-based column numbers -> field
_COLS = {
    "name": 3,
    "main_partner": 4,
    "on_website": 5,
    "client_type_raw": 6,
    "service_raw": 7,
    "impact_area_raw": 8,
    "topics_raw": 9,
    "objective": 10,
    "short_description": 11,
    "stage": 12,
    "notes": 13,
    "last_update": 14,
    "web_copy": 15,
    "impact_story": 17,
    "client_contact": 18,
}


@dataclass
class StagedRow:
    row_index: int
    name: str
    main_partner: str | None = None
    on_website: bool | None = None
    client_type_raw: str | None = None
    service_raw: str | None = None
    impact_area_raw: str | None = None
    topics_raw: str | None = None
    objective: str | None = None
    short_description: str | None = None
    stage: str | None = None
    notes: str | None = None
    last_update: str | None = None
    web_copy: str | None = None
    impact_story: str | None = None
    client_contact: str | None = None
    is_old_project: bool = False


def _text(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if s.endswith(".0"):  # openpyxl reads bare years/ints as floats
        s = s[:-2]
    return s or None


def _yesno(value: object) -> bool | None:
    s = _text(value)
    if s is None:
        return None
    return s.lower().startswith("y")


def parse_overview_xlsx(content: bytes) -> list[StagedRow]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[_SHEET] if _SHEET in wb.sheetnames else wb.active
    rows: list[StagedRow] = []
    old_mode = False
    for idx, cells in enumerate(
        ws.iter_rows(min_row=_HEADER_ROW + 1, values_only=True), start=_HEADER_ROW + 1
    ):

        def col(field: str, _row: tuple = cells) -> object:
            i = _COLS[field] - 1
            return _row[i] if i < len(_row) else None

        name = _text(col("name"))
        if name is None:
            continue
        if _SEPARATOR in name.lower():
            old_mode = True
            continue
        rows.append(
            StagedRow(
                row_index=idx,
                name=name,
                main_partner=_text(col("main_partner")),
                on_website=_yesno(col("on_website")),
                client_type_raw=_text(col("client_type_raw")),
                service_raw=_text(col("service_raw")),
                impact_area_raw=_text(col("impact_area_raw")),
                topics_raw=_text(col("topics_raw")),
                objective=_text(col("objective")),
                short_description=_text(col("short_description")),
                stage=_text(col("stage")),
                notes=_text(col("notes")),
                last_update=_text(col("last_update")),
                web_copy=_text(col("web_copy")),
                impact_story=_text(col("impact_story")),
                client_contact=_text(col("client_contact")),
                is_old_project=old_mode,
            )
        )
    wb.close()
    return rows


async def replace_staging(
    db: AsyncSession, batch_id: UUID, rows: list[StagedRow]
) -> tuple[int, int]:
    await db.execute(delete(PortfolioOverviewStagingDB))
    old = 0
    for r in rows:
        db.add(
            PortfolioOverviewStagingDB(
                import_batch=batch_id,
                row_index=r.row_index,
                name=r.name,
                main_partner=r.main_partner,
                on_website=r.on_website,
                client_type_raw=r.client_type_raw,
                service_raw=r.service_raw,
                impact_area_raw=r.impact_area_raw,
                topics_raw=r.topics_raw,
                objective=r.objective,
                short_description=r.short_description,
                stage=r.stage,
                notes=r.notes,
                last_update=r.last_update,
                web_copy=r.web_copy,
                impact_story=r.impact_story,
                client_contact=r.client_contact,
                is_old_project=r.is_old_project,
            )
        )
        old += 1 if r.is_old_project else 0
    await db.flush()
    logger.info("portfolio_overview_uploaded", batch_id=str(batch_id), rows=len(rows), old=old)
    return len(rows), old


STRONG = 0.85
CANDIDATE_THRESHOLD = 0.35


@dataclass
class ProjectCandidate:
    id: UUID
    name: str
    score: float


@dataclass
class CurrentProgram:
    program_id: UUID | None
    name: str | None


@dataclass
class SuggestedProject:
    project_id: UUID | None
    score: float


@dataclass
class ProgramCandidate:
    id: UUID
    name: str
    score: float


@dataclass
class SuggestedProgram:
    program_id: UUID | None
    score: float


def _program_first_default(
    name: str, program_pairs: list[tuple[str, UUID]]
) -> tuple[ProgramAction, UUID | None, str | None]:
    """Program-first default: link the strong fuzzy program match, else create by name."""
    ranked = rank(name, program_pairs, limit=1, threshold=CANDIDATE_THRESHOLD)
    best = ranked[0] if ranked else None
    if best is not None and best.score >= STRONG:
        return ProgramAction.LINK, best.payload, None
    return ProgramAction.CREATE, None, name


@dataclass
class StagingMatchData:
    staging_id: UUID
    name: str
    is_old_project: bool
    client_type_raw: str | None
    service_raw: str | None
    impact_area_raw: str | None
    suggested_project: SuggestedProject
    project_candidates: list[ProjectCandidate]
    current_program: CurrentProgram
    program_candidates: list[ProgramCandidate]
    suggested_program: SuggestedProgram


async def build_matches(db: AsyncSession, batch_id: UUID) -> list[StagingMatchData]:
    projects = (
        await db.execute(
            select(ProjectDB.id, ProjectDB.name, ProjectDB.program_id).where(
                ProjectDB.is_billable.is_(True), ProjectDB.is_absence.is_(False)
            )
        )
    ).all()
    prog_names = dict((await db.execute(select(ProgramDB.id, ProgramDB.name))).all())
    prog_of = {p.id: p.program_id for p in projects}
    candidates = [(p.name, (p.id, p.name)) for p in projects]
    program_pairs = [(name, pid) for pid, name in prog_names.items()]
    staging = (
        (
            await db.execute(
                select(PortfolioOverviewStagingDB)
                .where(PortfolioOverviewStagingDB.import_batch == batch_id)
                .order_by(PortfolioOverviewStagingDB.row_index)
            )
        )
        .scalars()
        .all()
    )
    result: list[StagingMatchData] = []
    for row in staging:
        ranked = rank(row.name, candidates, limit=5, threshold=CANDIDATE_THRESHOLD)
        proj_cands = [
            ProjectCandidate(id=s.payload[0], name=s.payload[1], score=s.score) for s in ranked
        ]
        best = ranked[0] if ranked else None
        if best is not None and best.score >= STRONG:
            suggested = SuggestedProject(project_id=best.payload[0], score=best.score)
            prog_id = prog_of.get(best.payload[0])
            current = CurrentProgram(program_id=prog_id, name=prog_names.get(prog_id))
        else:
            suggested = SuggestedProject(project_id=None, score=0.0)
            current = CurrentProgram(program_id=None, name=None)
        prog_ranked = rank(row.name, program_pairs, limit=5, threshold=CANDIDATE_THRESHOLD)
        program_candidates = [
            ProgramCandidate(id=s.payload, name=prog_names[s.payload], score=s.score)
            for s in prog_ranked
        ]
        best_prog = prog_ranked[0] if prog_ranked else None
        suggested_program = (
            SuggestedProgram(program_id=best_prog.payload, score=best_prog.score)
            if best_prog is not None and best_prog.score >= STRONG
            else SuggestedProgram(program_id=None, score=0.0)
        )
        result.append(
            StagingMatchData(
                staging_id=row.id,
                name=row.name,
                is_old_project=row.is_old_project,
                client_type_raw=row.client_type_raw,
                service_raw=row.service_raw,
                impact_area_raw=row.impact_area_raw,
                suggested_project=suggested,
                project_candidates=proj_cands,
                current_program=current,
                program_candidates=program_candidates,
                suggested_program=suggested_program,
            )
        )
    return result


async def seed_default_decisions(db: AsyncSession, batch_id: UUID) -> None:
    """Persist a program-first default decision per staged row (project_id NULL)."""
    prog_names = dict((await db.execute(select(ProgramDB.id, ProgramDB.name))).all())
    program_pairs = [(name, pid) for pid, name in prog_names.items()]
    rows = (
        (
            await db.execute(
                select(PortfolioOverviewStagingDB).where(
                    PortfolioOverviewStagingDB.import_batch == batch_id
                )
            )
        )
        .scalars()
        .all()
    )
    for row in rows:
        action, program_id, new_name = _program_first_default(row.name, program_pairs)
        row.program_action = action
        row.matched_program_id = program_id
        row.matched_project_id = None
        row.new_program_name = new_name
    await db.flush()


@dataclass
class CurrentBatch:
    batch_id: UUID
    row_count: int


async def get_current_batch(db: AsyncSession) -> CurrentBatch | None:
    """The batch that still has unapplied rows (single-batch invariant), or None."""
    row = (
        await db.execute(
            select(
                PortfolioOverviewStagingDB.import_batch,
                func.count(),
            )
            .where(PortfolioOverviewStagingDB.applied_at.is_(None))
            .group_by(PortfolioOverviewStagingDB.import_batch)
        )
    ).first()
    if row is None:
        return None
    return CurrentBatch(batch_id=row[0], row_count=row[1])


_TAXONOMY_FIELD = {
    "client-type": "client_type_raw",
    "service": "service_raw",
    "impact-area": "impact_area_raw",
    "topics": "topics_raw",
}
_OPEN_TAXONOMIES = {"topics"}


@dataclass
class DecisionInput:
    staging_id: UUID
    project_id: UUID | None
    program_action: ProgramAction
    program_id: UUID | None = None
    new_program_name: str | None = None


@dataclass
class ApplyResult:
    applied: int
    programs_created: int
    projects_linked_to_program: int
    programs_annotated: int
    skipped: int
    unmapped_terms: list[str]
    unresolved_clients: list[str]


async def _unique_program_name(db: AsyncSession, base: str) -> str:
    name = base.strip()[:255]
    suffix = 2
    while (
        await db.execute(select(ProgramDB.id).where(ProgramDB.name == name))
    ).first() is not None:
        name = f"{base.strip()[:248]} ({suffix})"
        suffix += 1
    return name


async def _resolve_program(
    db: AsyncSession, proj: ProjectDB, d: DecisionInput
) -> tuple[bool, bool]:
    """Returns (program_created, project_linked)."""
    if d.program_action == ProgramAction.INHERIT or d.program_action == ProgramAction.NONE:
        return False, False
    if d.program_action == ProgramAction.LINK:
        prog = (
            await db.execute(select(ProgramDB).where(ProgramDB.id == d.program_id))
        ).scalar_one()
        linked = proj.program_id != prog.id
        proj.program_id = prog.id
        await db.flush()
        return False, linked
    # CREATE
    prog = ProgramDB(name=await _unique_program_name(db, d.new_program_name or proj.name))
    db.add(prog)
    await db.flush()
    proj.program_id = prog.id
    await db.flush()
    return True, True


async def _upsert_profile(
    db: AsyncSession,
    row: PortfolioOverviewStagingDB,
    *,
    project_id: UUID | None = None,
    program_id: UUID | None = None,
) -> None:
    if project_id is not None:
        stmt = select(PortfolioProfileDB).where(PortfolioProfileDB.project_id == project_id)
    else:
        stmt = select(PortfolioProfileDB).where(PortfolioProfileDB.program_id == program_id)
    profile = (await db.execute(stmt)).scalar_one_or_none()
    if profile is None:
        profile = PortfolioProfileDB(project_id=project_id, program_id=program_id)
        db.add(profile)
    profile.objective = row.objective
    profile.short_description = row.short_description
    profile.web_copy = row.web_copy
    profile.impact_story = row.impact_story
    profile.stage = row.stage
    profile.main_partner = row.main_partner
    profile.on_website = bool(row.on_website)
    profile.source_batch = row.import_batch
    await db.flush()


async def _apply_terms(
    db: AsyncSession,
    row: PortfolioOverviewStagingDB,
    user_id: str | None,
    *,
    project_id: UUID | None = None,
    program_id: UUID | None = None,
) -> list[str]:
    unmapped: list[str] = []
    taxonomies = (await db.execute(select(TaxonomyDB))).scalars().all()
    by_slug = {t.slug: t for t in taxonomies}
    for slug, field in _TAXONOMY_FIELD.items():
        tax = by_slug.get(slug)
        raw = getattr(row, field)
        if tax is None or not raw:
            continue
        del_cond = (
            EntityTermDB.project_id == project_id
            if project_id is not None
            else EntityTermDB.program_id == program_id
        )
        await db.execute(delete(EntityTermDB).where(del_cond, EntityTermDB.taxonomy_id == tax.id))
        values = [v.strip() for v in raw.split(",") if v.strip()]
        if tax.cardinality == "single":
            values = values[:1]
        first = True
        for value in values:
            term = (
                await db.execute(
                    select(TaxonomyTermDB).where(
                        TaxonomyTermDB.taxonomy_id == tax.id,
                        func.lower(TaxonomyTermDB.name) == value.lower(),
                    )
                )
            ).scalar_one_or_none()
            if term is None:
                if slug in _OPEN_TAXONOMIES:
                    term = TaxonomyTermDB(taxonomy_id=tax.id, slug=slugify(value), name=value)
                    db.add(term)
                    await db.flush()
                else:
                    unmapped.append(f"{slug}: {value}")
                    continue
            db.add(
                EntityTermDB(
                    term_id=term.id,
                    taxonomy_id=tax.id,
                    project_id=project_id,
                    program_id=program_id,
                    is_primary=first and tax.allows_primary,
                    assigned_by=UUID(user_id) if user_id else None,
                )
            )
            first = False
        await db.flush()
    return unmapped


async def _cascade_client(
    db: AsyncSession, row: PortfolioOverviewStagingDB, proj: ProjectDB
) -> str | None:
    if not row.main_partner:
        return None
    clients = (await db.execute(select(ClientDB.id, ClientDB.name))).all()
    ranked = rank(row.main_partner, [(c.name, c.id) for c in clients], limit=1, threshold=0.9)
    if not ranked:
        return row.main_partner
    client_id = ranked[0].payload
    # matched project + (if it has a program) its siblings, only where NULL
    cond = ProjectDB.client_id.is_(None)
    if proj.program_id is not None:
        await db.execute(
            ProjectDB.__table__.update()
            .where(ProjectDB.program_id == proj.program_id, cond)
            .values(client_id=client_id)
        )
    else:
        await db.execute(
            ProjectDB.__table__.update()
            .where(ProjectDB.id == proj.id, cond)
            .values(client_id=client_id)
        )
    if row.client_contact:
        client = (await db.execute(select(ClientDB).where(ClientDB.id == client_id))).scalar_one()
        if not client.primary_contact:
            client.primary_contact = row.client_contact[:255]
    await db.flush()
    return None


async def _cascade_client_to_program(
    db: AsyncSession, row: PortfolioOverviewStagingDB, program_id: UUID
) -> str | None:
    if not row.main_partner:
        return None
    clients = (await db.execute(select(ClientDB.id, ClientDB.name))).all()
    ranked = rank(row.main_partner, [(c.name, c.id) for c in clients], limit=1, threshold=0.9)
    if not ranked:
        return row.main_partner
    client_id = ranked[0].payload
    await db.execute(
        ProjectDB.__table__.update()
        .where(ProjectDB.program_id == program_id, ProjectDB.client_id.is_(None))
        .values(client_id=client_id)
    )
    if row.client_contact:
        client = (await db.execute(select(ClientDB).where(ClientDB.id == client_id))).scalar_one()
        if not client.primary_contact:
            client.primary_contact = row.client_contact[:255]
    await db.flush()
    return None


async def apply_decisions(
    db: AsyncSession, batch_id: UUID, decisions: list[DecisionInput], user_id: str | None
) -> ApplyResult:
    applied = created = linked = annotated = skipped = 0
    unmapped: list[str] = []
    unresolved: list[str] = []
    now = datetime.now(UTC)
    for d in decisions:
        staging_id_str = str(d.staging_id)
        row = (
            await db.execute(
                select(PortfolioOverviewStagingDB).where(
                    PortfolioOverviewStagingDB.id == d.staging_id,
                    PortfolioOverviewStagingDB.import_batch == batch_id,
                )
            )
        ).scalar_one_or_none()
        if row is None:
            continue
        if d.project_id is None and d.program_action not in (
            ProgramAction.LINK,
            ProgramAction.CREATE,
        ):
            row.program_action = ProgramAction.NONE
            row.decided_by = UUID(user_id) if user_id else None
            row.decided_at = now
            await db.flush()
            skipped += 1
            continue
        row_unmapped: list[str] = []
        row_unresolved: str | None = None
        did_create = did_link = is_program_anchor = False
        try:
            async with db.begin_nested():
                if d.project_id is not None:
                    proj = (
                        await db.execute(select(ProjectDB).where(ProjectDB.id == d.project_id))
                    ).scalar_one()
                    did_create, did_link = await _resolve_program(db, proj, d)
                    await _upsert_profile(db, row, project_id=proj.id)
                    row_unmapped = await _apply_terms(db, row, user_id, project_id=proj.id)
                    row_unresolved = await _cascade_client(db, row, proj)
                    row.matched_project_id = proj.id
                    row.matched_program_id = proj.program_id
                else:
                    is_program_anchor = True
                    if d.program_action == ProgramAction.CREATE:
                        prog = ProgramDB(
                            name=await _unique_program_name(db, d.new_program_name or row.name)
                        )
                        db.add(prog)
                        await db.flush()
                        did_create = True
                    else:  # LINK
                        prog = (
                            await db.execute(select(ProgramDB).where(ProgramDB.id == d.program_id))
                        ).scalar_one()
                    await _upsert_profile(db, row, program_id=prog.id)
                    row_unmapped = await _apply_terms(db, row, user_id, program_id=prog.id)
                    row_unresolved = await _cascade_client_to_program(db, row, prog.id)
                    row.matched_project_id = None
                    row.matched_program_id = prog.id
                row.program_action = d.program_action
                row.decided_by = UUID(user_id) if user_id else None
                row.decided_at = now
        except Exception:
            logger.warning("portfolio_overview_row_failed", staging_id=staging_id_str)
            continue
        unmapped.extend(row_unmapped)
        if row_unresolved:
            unresolved.append(row_unresolved)
        applied += 1
        created += 1 if did_create else 0
        linked += 1 if did_link else 0
        annotated += 1 if is_program_anchor else 0
    logger.info(
        "portfolio_overview_applied",
        batch_id=str(batch_id),
        applied=applied,
        created=created,
        linked=linked,
        annotated=annotated,
        skipped=skipped,
    )
    return ApplyResult(
        applied=applied,
        programs_created=created,
        projects_linked_to_program=linked,
        programs_annotated=annotated,
        skipped=skipped,
        unmapped_terms=unmapped,
        unresolved_clients=unresolved,
    )
