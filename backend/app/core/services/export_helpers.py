"""Shared XLSX formatting helpers for export modules."""

import calendar
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Soft palette — same hues (#4CAF50, #FFC107, #F44336) at two opacity levels on white.
# Strong (~35% opacity): scores and dimensions (level 0-1)
GREEN_FILL = PatternFill(start_color="C1E3C2", end_color="C1E3C2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFE9A8", end_color="FFE9A8", fill_type="solid")
RED_FILL = PatternFill(start_color="FBBDB9", end_color="FBBDB9", fill_type="solid")

# Subtle (~15% opacity): indicators (level 2)
GREEN_FILL_SUBTLE = PatternFill(start_color="E4F3E5", end_color="E4F3E5", fill_type="solid")
YELLOW_FILL_SUBTLE = PatternFill(start_color="FFF6DA", end_color="FFF6DA", fill_type="solid")
RED_FILL_SUBTLE = PatternFill(start_color="FDE3E1", end_color="FDE3E1", fill_type="solid")

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DIM_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
DIM_FONT = Font(bold=True, size=11)

SCORE_FONT = Font(bold=True, size=12)
SCORE_FILL = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

DEFAULT_GREEN_THRESHOLD = 80
DEFAULT_YELLOW_THRESHOLD = 60


def apply_score_traffic_light(
    cell,
    value: float | int | None,
    green: float = DEFAULT_GREEN_THRESHOLD,
    yellow: float = DEFAULT_YELLOW_THRESHOLD,
) -> None:
    """Apply strong green/yellow/red fill for scores and dimensions (0-100 scale)."""
    if value is None:
        return
    if value >= green:
        cell.fill = GREEN_FILL
    elif value >= yellow:
        cell.fill = YELLOW_FILL
    else:
        cell.fill = RED_FILL


def apply_indicator_traffic_light(
    cell,
    value: float | int | None,
    green: float = 0.80,
    yellow: float = 0.60,
) -> None:
    """Apply subtle green/yellow/red fill for indicators (0-1 scale)."""
    if value is None:
        return
    if value >= green:
        cell.fill = GREEN_FILL_SUBTLE
    elif value >= yellow:
        cell.fill = YELLOW_FILL_SUBTLE
    else:
        cell.fill = RED_FILL_SUBTLE


def apply_header_style(ws: Worksheet, row: int = 1) -> None:
    """Apply dark header style to a row."""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_row_style(ws: Worksheet, row: int, level: int) -> None:
    """Apply style based on hierarchy level."""
    for cell in ws[row]:
        cell.border = THIN_BORDER
        if level == 0:
            cell.font = SCORE_FONT
            cell.fill = SCORE_FILL
        elif level == 1:
            cell.font = DIM_FONT
            cell.fill = DIM_FILL


def format_month_header(year: int, month: int) -> str:
    """Format year/month as 'Jan 2025'."""
    return f"{calendar.month_abbr[month]} {year}"


def set_column_widths(ws: Worksheet, widths: dict[str, int]) -> None:
    """Set column widths from a dict of column letter -> width."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def freeze_panes(ws: Worksheet, row: int, col: int) -> None:
    """Freeze rows above and columns to the left of the given cell."""
    ws.freeze_panes = f"{get_column_letter(col)}{row}"


def save_to_bytes(wb: Workbook) -> BytesIO:
    """Save workbook to an in-memory BytesIO buffer."""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
