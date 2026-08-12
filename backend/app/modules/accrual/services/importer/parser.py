"""Excel spreadsheet parser for the CEO's accrual workbook.

Pure parsing — no DB I/O. Returns ``SpreadsheetRow`` instances that downstream
phases (snapshot, resolve, render) consume.
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class SpreadsheetRow:
    type: str
    code: str | None
    pm: str | None
    name: str | None
    value: Decimal
    rate: Decimal
    value_eur: Decimal
    start_date: date | None
    end_date: date | None
    duration: int | None
    monthly: dict[tuple[int, int], Decimal] = field(default_factory=dict)


_SUFFIX_PATTERN = re.compile(r"\.[\w/]+$")


def _normalize_code(code: str | None) -> str | None:
    """Strip ALL whitespace + collapse repeated dots + upper-case.

    Handles divergences such as ``"LSE .TPI2025 .35054413"`` (internal spaces),
    ``"ICIMOD..34229341"`` (doubled dots), and trailing whitespace from copy-
    paste mistakes between Excel and VizzHub.
    """
    if not code:
        return None
    no_ws = "".join(code.split())
    collapsed = re.sub(r"\.{2,}", ".", no_ws)
    return collapsed.upper() or None


def _code_prefix(code: str | None) -> str | None:
    """Return the code with its trailing dot-segment stripped, or None.

    Trailing segments are typically Jira issue IDs (``.32232147``), signing
    years (``.24``), or version markers (``.24/25``, ``.2``). Returns None
    when the normalised code has no dot or no trailing alphanumeric segment.
    Only returns a *changed* form (no-op when stripping wouldn't shorten).
    """
    norm = _normalize_code(code)
    if not norm:
        return None
    stripped = _SUFFIX_PATTERN.sub("", norm)
    return stripped if stripped and stripped != norm else None


def _coerce_date(v: object) -> date | None:
    if v is None:
        return None
    return v.date() if hasattr(v, "date") else v  # type: ignore[return-value]


def _month_columns(ws: Worksheet) -> dict[int, tuple[int, int]]:
    """Map column index -> (year, month) from header rows 5/6 (monthly cols start at 13)."""
    columns: dict[int, tuple[int, int]] = {}
    for c in range(13, ws.max_column + 1):
        y = ws.cell(row=5, column=c).value
        m = ws.cell(row=6, column=c).value
        if isinstance(y, (int, float)) and isinstance(m, (int, float)):
            columns[c] = (int(y), int(m))
    return columns


def _monthly_amounts(
    ws: Worksheet, r: int, columns: dict[int, tuple[int, int]]
) -> dict[tuple[int, int], Decimal]:
    monthly: dict[tuple[int, int], Decimal] = {}
    for c, (y, m) in columns.items():
        v = ws.cell(row=r, column=c).value
        if isinstance(v, (int, float)):
            monthly[(y, m)] = Decimal(str(v)).quantize(Decimal("0.01"))
    return monthly


def _parse_data_row(
    ws: Worksheet, r: int, columns: dict[int, tuple[int, int]]
) -> SpreadsheetRow | None:
    """Parse one data row; None when the row is blank or has no rate."""
    type_v = ws.cell(row=r, column=1).value
    if not type_v:
        return None
    rate_v = ws.cell(row=r, column=7).value
    if rate_v is None:
        return None

    code_v = ws.cell(row=r, column=3).value
    return SpreadsheetRow(
        type=str(type_v),
        code=str(code_v) if code_v else None,
        pm=ws.cell(row=r, column=4).value,
        name=ws.cell(row=r, column=5).value,
        value=Decimal(str(ws.cell(row=r, column=6).value or 0)),
        rate=Decimal(str(rate_v)),
        value_eur=Decimal(str(ws.cell(row=r, column=8).value or 0)),
        start_date=_coerce_date(ws.cell(row=r, column=9).value),
        end_date=_coerce_date(ws.cell(row=r, column=10).value),
        duration=ws.cell(row=r, column=11).value,
        monthly=_monthly_amounts(ws, r, columns),
    )


def parse_spreadsheet(path: Path) -> list[SpreadsheetRow]:
    """Parse the CEO's accrual workbook into structured rows.

    Sheet layout (per inspection 2026-05-22): sheet 'Sales', year row 5,
    month/header row 6, data rows from row 7. Monthly columns start at 13.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb["Sales"]
    columns = _month_columns(ws)

    rows: list[SpreadsheetRow] = []
    for r in range(7, ws.max_row + 1):
        row = _parse_data_row(ws, r, columns)
        if row is not None:
            rows.append(row)
    return rows


def consolidate_duplicate_rows(rows: list[SpreadsheetRow]) -> list[SpreadsheetRow]:
    """Merge Excel rows sharing the same normalized code into one synthetic row.

    The CEO's spreadsheet uses multiple rows per code to track contract
    amendments / extensions (e.g. EVF.ESAGDA appears 4 times, one per annual
    renewal). The accrual model has one project per code, so for each duplicate
    group we sum ``value`` and merge ``monthly`` (summing if both rows cover
    the same month). ``start_date``/``end_date`` are set to the min/max across
    the group so downstream date-extension logic can spot range mismatches.

    Rows with no code are passed through untouched. The rate of the first row
    in the group is kept on the synthetic row (rate aggregation per period is
    handled in ``bootstrap_periods`` against the raw rows, not these merged
    ones).
    """
    by_code: dict[str | None, list[SpreadsheetRow]] = defaultdict(list)
    for r in rows:
        key = _normalize_code(r.code)
        by_code[key].append(r)

    consolidated: list[SpreadsheetRow] = []
    for key, group in by_code.items():
        if key is None or len(group) == 1:
            consolidated.extend(group)
        else:
            consolidated.append(_merge_group(group))
    return consolidated


def _merge_group(group: list[SpreadsheetRow]) -> SpreadsheetRow:
    """Merge a duplicate-code group into one synthetic row (see consolidate_duplicate_rows)."""
    monthly: dict[tuple[int, int], Decimal] = defaultdict(lambda: Decimal("0"))
    for r in group:
        for ym, amount in r.monthly.items():
            monthly[ym] += amount
    starts = [r.start_date for r in group if r.start_date]
    ends = [r.end_date for r in group if r.end_date]
    head = group[0]
    return SpreadsheetRow(
        type=head.type,
        code=head.code,
        pm=head.pm,
        name=(head.name or "") + f" (+{len(group) - 1} amendment{'s' if len(group) > 2 else ''})",
        value=sum((r.value for r in group), Decimal("0")),
        rate=head.rate,
        value_eur=sum((r.value_eur for r in group), Decimal("0")),
        start_date=min(starts) if starts else None,
        end_date=max(ends) if ends else None,
        duration=None,
        monthly=dict(monthly),
    )
