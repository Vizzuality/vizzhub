"""Scorecard-specific XLSX helpers — methodology sheet and config lookups."""

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.config import ScoringConfig
from app.core.services.export_helpers import (
    GREEN_FILL,
    GREEN_FILL_SUBTLE,
    YELLOW_FILL,
    YELLOW_FILL_SUBTLE,
    RED_FILL,
    RED_FILL_SUBTLE,
    apply_header_style,
    set_column_widths,
)
from app.modules.scorecard.services.export_definitions import (
    DIMENSION_DEFINITIONS,
    INDICATOR_DEFINITIONS,
)


DEFAULT_GREEN_THRESHOLD = 80
DEFAULT_YELLOW_THRESHOLD = 60


def create_methodology_sheet(wb: Workbook, config: ScoringConfig) -> Worksheet:
    """Create the Methodology sheet with scoring model explanation and current config."""
    ws = wb.create_sheet("Methodology")

    green_th = _get_threshold(config, "threshold_green", DEFAULT_GREEN_THRESHOLD)
    yellow_th = _get_threshold(config, "threshold_yellow", DEFAULT_YELLOW_THRESHOLD)

    ws.append(["Scoring Methodology"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Scoring Model"])
    ws.cell(row=3, column=1).font = Font(bold=True, size=12)
    ws.append(["Raw Metrics", "->", "Normalized Indicators (0-1)", "->", "Weighted Scores (0-100)"])
    ws.append(["Collectors fetch data from Jira and GitHub. Manual inputs supplement automated data."])
    ws.append(["Indicators are normalized to a 0-1 scale. Scores are weighted averages scaled to 0-100."])
    ws.append([])

    ws.append(["Traffic Light Legend"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    ind_green = green_th / 100
    ind_yellow = yellow_th / 100

    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="Scores & Dimensions (0-100)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)
    ws.cell(row=row, column=2, value="Indicators (0-1)")
    ws.cell(row=row, column=2).font = Font(bold=True, size=10)
    ws.cell(row=row, column=3, value="Category")
    ws.cell(row=row, column=3).font = Font(bold=True, size=10)
    row += 1
    ws.cell(row=row, column=1, value=f"Score >= {green_th:.0f}").fill = GREEN_FILL
    ws.cell(row=row, column=2, value=f"Value >= {ind_green:.2f}").fill = GREEN_FILL_SUBTLE
    ws.cell(row=row, column=3, value="Good performance")
    row += 1
    ws.cell(row=row, column=1, value=f"Score >= {yellow_th:.0f}").fill = YELLOW_FILL
    ws.cell(row=row, column=2, value=f"Value >= {ind_yellow:.2f}").fill = YELLOW_FILL_SUBTLE
    ws.cell(row=row, column=3, value="At risk / needs attention")
    row += 1
    ws.cell(row=row, column=1, value=f"Score < {yellow_th:.0f}").fill = RED_FILL
    ws.cell(row=row, column=2, value=f"Value < {ind_yellow:.2f}").fill = RED_FILL_SUBTLE
    ws.cell(row=row, column=3, value="Critical / poor performance")
    ws.append([])

    ws.append(["Snapshot Types"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Cumulative", "Project-to-date metrics from start_date to period end"])
    ws.append(["Punctual", "Single month metrics for that period only"])
    ws.append([])

    ws.append(["Global Dimension Weights"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Dimension", "Weight"])
    dims = ["time", "cost", "quality", "value", "satisfaction", "flow", "engineering", "risk"]
    for dim in dims:
        weight = config.get_weight("global", dim)
        ws.append([f"P_{dim}", f"{weight:.0%}"])
    ws.append([])

    ws.append(["KPI Reference"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    headers = ["Dimension", "Indicator", "Description", "Formula", "Target", "Weight"]
    ws.append(headers)
    header_row = ws.max_row
    apply_header_style(ws, header_row)

    for dim_def in DIMENSION_DEFINITIONS:
        dim_key = dim_def["key"].replace("p_", "")
        for ind_key in dim_def["indicators"]:
            ind = INDICATOR_DEFINITIONS[ind_key]
            target = _safe_get_target(config, ind_key)
            weight = _safe_get_weight(config, dim_key, ind_key)
            ws.append([
                dim_def["name"],
                ind["name"],
                ind["description"],
                ind["formula"],
                target,
                weight,
            ])

    set_column_widths(ws, {"A": 25, "B": 28, "C": 50, "D": 60, "E": 12, "F": 12})

    return ws


def _get_threshold(config: ScoringConfig, name: str, default: float) -> float:
    """Get a threshold constant from config, falling back to default."""
    try:
        val = config.get_constant(name)
        return val if val > 0 else default
    except (KeyError, ValueError):
        return default


def _safe_get_target(config: ScoringConfig, indicator_key: str) -> str:
    """Get target value as string, return '-' if not configured."""
    try:
        val = config.get_target(indicator_key)
        if val is not None and abs(val) < 1e-9:
            return "-"
        return str(val)
    except (KeyError, ValueError):
        return "-"


def _safe_get_weight(config: ScoringConfig, _dim_key: str, ind_key: str) -> str:
    """Get weight value as string, return '-' if not configured."""
    weight_map = {
        "spi": ("time", "spi"),
        "on_time_milestones": ("time", "milestones"),
        "cpi": ("cost", "cpi"),
        "cost_variance_pct": ("cost", "variance"),
        "defect_density": ("quality", "defect_density"),
        "escaped_rate": ("quality", "escaped_rate"),
        "mttr_hours": ("quality", "mttr"),
        "governance_compliance": ("quality", "governance"),
        "story_review_ratio": ("quality", "story_review"),
        "pr_review_ratio": ("quality", "pr_review"),
        "change_failure_rate": ("quality", "change_failure_rate"),
        "post_contract_tasks": ("quality", "post_contract_tasks"),
        "okr_impact": ("value", "okr_impact"),
        "pm_satisfaction": ("satisfaction", "pm_estimation"),
        "client_satisfaction": ("satisfaction", "client_survey"),
        "lead_time_days": ("flow", "lead_time"),
        "commitment_reliability": ("flow", "commitment_reliability"),
        "pr_size_median": ("flow", "pr_size"),
        "review_turnaround_hours": ("flow", "review_turnaround"),
        "deployment_frequency": ("flow", "deployment_frequency"),
        "test_maturity": ("engineering", "test_maturity"),
        "arch_checklist": ("engineering", "architecture"),
        "prs_without_review": ("risk", "pr_no_review"),
        "high_vulns": ("risk", "high_vulns"),
    }
    mapping = weight_map.get(ind_key)
    if not mapping:
        return "-"
    try:
        val = config.get_weight(mapping[0], mapping[1])
        return f"{val:.0%}"
    except (KeyError, ValueError):
        return "-"
