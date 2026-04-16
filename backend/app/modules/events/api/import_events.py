"""Excel import endpoint for bulk event creation."""

from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Annotated

import structlog
from fastapi import APIRouter, Depends, HTTPException, UploadFile
from sqlalchemy import func, select

from app.core.api.deps import DBSession
from app.core.auth import TokenData
from app.core.models.user import UserDB
from app.core.permissions import Action, require_permission
from app.core.sql_helpers import user_display_name_expr
from app.modules.events.models.event import EventDB
from app.modules.events.models.event_attendee import EventAttendeeDB

logger = structlog.get_logger()

router = APIRouter()

EventsManager = Annotated[TokenData, Depends(require_permission(Action.EVENTS_MANAGE))]

EXPECTED_HEADERS = {
    "event name": "name",
    "type": "event_type",
    "primary theme": "theme_primary",
    "secondary theme": "theme_secondary",
    "region focus": "region_focus",
    "city": "location_city",
    "country": "location_country",
    "start date": "start_date",
    "end date": "end_date",
    "cost": "cost",
    "rating": "rating",
    "url": "url",
    "observations": "observations",
    "attendee name": "attendee_name",
    "attendee role": "attendee_role",
}


def _build_user_lookup(users: list) -> dict[str, "UserDB"]:
    """Map lowercase display names to UserDB objects."""
    lookup: dict[str, UserDB] = {}
    for u in users:
        if u.first_name and u.last_name:
            key = f"{u.first_name} {u.last_name}".lower()
            lookup[key] = u
        if u.name:
            lookup[u.name.lower()] = u
    return lookup


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except (ValueError, TypeError):
        return None


def _parse_decimal(value) -> Decimal:
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _parse_rating(value) -> int | None:
    if value is None:
        return None
    try:
        r = int(value)
        return r if 1 <= r <= 5 else None
    except (ValueError, TypeError):
        return None


def _safe_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


@router.post("/import")
async def import_events_from_excel(
    file: UploadFile,
    db: DBSession,
    user: EventsManager,
) -> dict:
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed on the server",
        )

    content = await file.read()

    try:
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read Excel file")

    sheet_name = "Events" if "Events" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty spreadsheet")

    header_row = rows[0]
    col_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            normalized = str(cell).strip().lower()
            if normalized in EXPECTED_HEADERS:
                col_map[EXPECTED_HEADERS[normalized]] = idx

    if "name" not in col_map or "start_date" not in col_map:
        raise HTTPException(
            status_code=400,
            detail="Spreadsheet must have 'Event Name' and 'Start Date' columns",
        )

    def _cell(row, field: str):
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    # Group rows by (event_name, start_date) to deduplicate
    event_groups: dict[tuple[str, date], dict] = {}
    skipped_rows = 0

    for row_num, row in enumerate(rows[1:], start=2):
        event_name = _safe_str(_cell(row, "name"))
        start_date_val = _parse_date(_cell(row, "start_date"))

        if not event_name or not start_date_val:
            skipped_rows += 1
            continue

        key = (event_name, start_date_val)
        if key not in event_groups:
            event_groups[key] = {
                "name": event_name,
                "event_type": _safe_str(_cell(row, "event_type")) or "Other",
                "theme_primary": _safe_str(_cell(row, "theme_primary")) or "Other",
                "theme_secondary": _safe_str(_cell(row, "theme_secondary")),
                "region_focus": _safe_str(_cell(row, "region_focus")) or "Global",
                "location_city": _safe_str(_cell(row, "location_city")),
                "location_country": _safe_str(_cell(row, "location_country")),
                "start_date": start_date_val,
                "end_date": _parse_date(_cell(row, "end_date")),
                "cost": _parse_decimal(_cell(row, "cost")),
                "rating": _parse_rating(_cell(row, "rating")),
                "url": _safe_str(_cell(row, "url")),
                "observations": _safe_str(_cell(row, "observations")),
                "attendees": [],
            }

        attendee_name = _safe_str(_cell(row, "attendee_name"))
        attendee_role = _safe_str(_cell(row, "attendee_role"))
        if attendee_name:
            event_groups[key]["attendees"].append({
                "name": attendee_name,
                "role": attendee_role or "Attendee",
            })

    # Load all users for name matching
    user_result = await db.execute(select(UserDB).where(UserDB.active.is_(True)))
    all_users = user_result.scalars().all()
    user_lookup = _build_user_lookup(all_users)

    events_created = 0
    attendees_matched = 0
    unmatched_attendee_names: list[str] = []

    for group in event_groups.values():
        attendees_data = group.pop("attendees")

        event = EventDB(**group, created_by=user.user_id)
        db.add(event)
        await db.flush()
        events_created += 1

        seen_user_ids = set()
        for att in attendees_data:
            matched_user = user_lookup.get(att["name"].lower())
            if matched_user is None:
                if att["name"] not in unmatched_attendee_names:
                    unmatched_attendee_names.append(att["name"])
                continue

            if matched_user.id in seen_user_ids:
                continue
            seen_user_ids.add(matched_user.id)

            db.add(EventAttendeeDB(
                event_id=event.id,
                user_id=matched_user.id,
                role=att["role"],
            ))
            attendees_matched += 1

    await db.commit()
    wb.close()

    logger.info(
        "events_imported",
        events_created=events_created,
        attendees_matched=attendees_matched,
        unmatched=len(unmatched_attendee_names),
        skipped_rows=skipped_rows,
    )

    return {
        "events_created": events_created,
        "attendees_matched": attendees_matched,
        "unmatched_attendee_names": unmatched_attendee_names,
        "skipped_rows": skipped_rows,
    }
