"""XLSX export service for ISO access review data."""

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.core.services.export_helpers import (
    apply_header_style,
    save_to_bytes,
    set_column_widths,
)

DATETIME_FORMAT_UTC = "%Y-%m-%d %H:%M UTC"


class IsoExportService:
    """Generates XLSX exports for ISO access review snapshots."""

    def export_snapshots(
        self,
        snapshots_with_reviews: list[tuple[dict, dict | None, list[dict]]],
    ) -> BytesIO:
        wb = Workbook()

        for i, (snapshot, review, actions) in enumerate(snapshots_with_reviews):
            captured = snapshot["captured_at"]
            if isinstance(captured, datetime):
                tab_name = f"Review {captured.strftime('%Y-%m-%d')}"
            else:
                tab_name = f"Review {str(captured)[:10]}"
            existing = [s for s in wb.sheetnames if s.startswith(tab_name)]
            if existing:
                tab_name = f"{tab_name} ({len(existing)})"

            if i == 0:
                ws = wb.active
                ws.title = tab_name
            else:
                ws = wb.create_sheet(tab_name)
            self._write_snapshot_tab(ws, snapshot, review, actions)

        return save_to_bytes(wb)

    def _write_snapshot_tab(
        self,
        ws,
        snapshot: dict,
        review: dict | None,
        actions: list[dict],
    ) -> None:
        provider = snapshot.get("provider", "google_workspace")

        self._write_iso_header(ws, snapshot, review)
        ws.append([])

        if review and review.get("diff_summary"):
            self._write_diff_summary(ws, review["diff_summary"])
            ws.append([])

        if actions:
            self._write_actions_table(ws, actions)
            ws.append([])

        data = snapshot.get("data", {})

        if provider == "github":
            self._write_github_data_tables(ws, data)
        elif provider == "jira":
            self._write_jira_data_tables(ws, data)
        else:
            self._write_gw_data_tables(ws, data)

        set_column_widths(
            ws, {"A": 25, "B": 30, "C": 20, "D": 20, "E": 18, "F": 18, "G": 16}
        )

    def _write_gw_data_tables(self, ws, data: dict) -> None:
        self._write_users_table(ws, data.get("users", []))
        ws.append([])
        self._write_groups_table(
            ws, data.get("groups", []), data.get("group_members", {})
        )
        ws.append([])
        self._write_group_members_table(ws, data.get("group_members", {}))
        ws.append([])
        self._write_admins_table(ws, data.get("role_assignments", []))

    def _write_github_data_tables(self, ws, data: dict) -> None:
        self._write_github_members_table(ws, data.get("members", []))
        ws.append([])
        self._write_github_teams_table(
            ws, data.get("teams", []), data.get("team_members", {})
        )
        ws.append([])
        self._write_github_team_members_table(ws, data.get("team_members", {}))
        ws.append([])
        self._write_github_outside_collaborators_table(
            ws, data.get("outside_collaborators", [])
        )

    def _write_iso_header(self, ws, snapshot: dict, review: dict | None) -> None:
        provider = snapshot.get("provider", "google_workspace")
        metadata = snapshot.get("source_metadata", {})
        if provider == "github":
            org_label = metadata.get("org", "")
        elif provider == "jira":
            org_label = metadata.get("site_url", "")
        else:
            org_label = metadata.get("domain", "")

        captured = snapshot["captured_at"]
        captured_str = (
            captured.strftime(DATETIME_FORMAT_UTC)
            if isinstance(captured, datetime)
            else str(captured)
        )

        summary = snapshot.get("summary", {})
        admins_row = ("Total Admins", summary.get("total_admins", 0))

        if provider == "github":
            summary_rows = [
                ("Total Members", summary.get("total_members", 0)),
                admins_row,
                ("Total Teams", summary.get("total_teams", 0)),
                ("Outside Collaborators", summary.get("outside_collaborators", 0)),
            ]
        elif provider == "jira":
            summary_rows = [
                ("Total Users", summary.get("total_users", 0)),
                admins_row,
                ("Total Groups", summary.get("total_groups", 0)),
                ("External Users", summary.get("external_users", 0)),
            ]
        else:
            summary_rows = [
                ("Total Users", summary.get("total_users", 0)),
                admins_row,
                ("Total Groups", summary.get("total_groups", 0)),
                ("External Members", summary.get("external_members", 0)),
            ]

        header_rows = [
            ("Organization", org_label),
            ("Provider", provider),
            ("Snapshot Date", captured_str),
            *summary_rows,
            ("Review Scope", review["scope"] if review else "N/A"),
            ("Reviewer", review.get("reviewer_email", "") if review else ""),
            ("Status", review["status"] if review else "No review"),
            ("Signed By", review.get("signed_by_email", "") if review else ""),
            ("Signed Date", self._format_signed_date(review)),
            ("Notes", review.get("notes", "") if review else ""),
            ("Export Date", datetime.now(timezone.utc).strftime(DATETIME_FORMAT_UTC)),
        ]

        for label, value in header_rows:
            ws.append([label, value])
            row = ws.max_row
            ws.cell(row=row, column=1).font = Font(bold=True)

    @staticmethod
    def _format_signed_date(review: dict | None) -> str:
        if not review or not review.get("signed_at"):
            return ""
        signed_at = review["signed_at"]
        if isinstance(signed_at, datetime):
            return signed_at.strftime(DATETIME_FORMAT_UTC)
        return str(signed_at)

    def _write_diff_summary(self, ws, diff_summary: dict) -> None:
        ws.append(["Diff Summary"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Change Type", "Count"])
        apply_header_style(ws, ws.max_row)

        mapping = [
            ("New Users", "new_user"),
            ("Removed Users", "removed_user"),
            ("Role Changes", "role_change"),
            ("New External", "new_external"),
            ("Removed External", "removed_external"),
            ("Group Changes", "group_membership_change"),
        ]
        for label, key in mapping:
            ws.append([label, diff_summary.get(key, 0)])

    def _write_actions_table(self, ws, actions: list[dict]) -> None:
        ws.append(["Actions"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        headers = [
            "Subject",
            "Type",
            "Change Type",
            "Details",
            "Action Taken",
            "Justification",
            "Exception Until",
        ]
        ws.append(headers)
        apply_header_style(ws, ws.max_row)

        for action in actions:
            details = ""
            prev = action.get("previous_value")
            curr = action.get("current_value")
            if prev:
                details += f"Previous: {prev}"
            if curr:
                if details:
                    details += " | "
                details += f"Current: {curr}"

            ws.append(
                [
                    action.get("subject_label", action.get("subject_id", "")),
                    action.get("subject_type", ""),
                    action.get("change_type", ""),
                    details or "\u2014",
                    action.get("action_taken", ""),
                    action.get("justification", ""),
                    (
                        str(action["exception_until"])
                        if action.get("exception_until")
                        else ""
                    ),
                ]
            )

    def _write_users_table(self, ws, users: list[dict]) -> None:
        ws.append(["Users"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Name", "Email", "Status", "Org Unit"])
        apply_header_style(ws, ws.max_row)

        for user in users:
            ws.append(
                [
                    user.get("name", ""),
                    user.get("email", ""),
                    "Suspended" if user.get("suspended") else "Active",
                    user.get("org_unit_path", ""),
                ]
            )

    def _write_groups_table(
        self,
        ws,
        groups: list[dict],
        group_members: dict[str, list],
    ) -> None:
        ws.append(["Groups"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Name", "Email", "Members"])
        apply_header_style(ws, ws.max_row)

        for group in groups:
            email = group.get("email", "")
            members = group_members.get(email, [])
            ws.append(
                [
                    group.get("name", ""),
                    email,
                    len(members),
                ]
            )

    def _write_group_members_table(self, ws, group_members: dict[str, list]) -> None:
        ws.append(["Group Members"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Group Email", "Member Email", "Role", "Type"])
        apply_header_style(ws, ws.max_row)

        for group_email in sorted(group_members.keys()):
            for member in group_members[group_email]:
                ws.append(
                    [
                        group_email,
                        member.get("email", ""),
                        member.get("role", ""),
                        member.get("type", ""),
                    ]
                )

    def _write_admins_table(self, ws, role_assignments: list[dict]) -> None:
        ws.append(["Admins"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Email", "Role Name"])
        apply_header_style(ws, ws.max_row)

        for ra in role_assignments:
            ws.append(
                [
                    ra.get("user_email", ""),
                    ra.get("role_name", ""),
                ]
            )

    # --- GitHub-specific tables ---

    def _write_github_members_table(self, ws, members: list[dict]) -> None:
        ws.append(["Members"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Login", "Name", "Email", "Role"])
        apply_header_style(ws, ws.max_row)

        for m in members:
            ws.append([
                m.get("login", ""),
                m.get("name", "") or "",
                m.get("email", "") or "",
                m.get("role", ""),
            ])

    def _write_github_teams_table(
        self, ws, teams: list[dict], team_members: dict[str, list]
    ) -> None:
        ws.append(["Teams"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Name", "Slug", "Parent", "Privacy", "Members"])
        apply_header_style(ws, ws.max_row)

        for t in teams:
            slug = t.get("slug", "")
            members = team_members.get(slug, [])
            ws.append([
                t.get("name", ""),
                slug,
                t.get("parent_slug", "") or "",
                t.get("privacy", ""),
                len(members),
            ])

    def _write_github_team_members_table(
        self, ws, team_members: dict[str, list]
    ) -> None:
        ws.append(["Team Members"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Team", "Login", "Role"])
        apply_header_style(ws, ws.max_row)

        for slug in sorted(team_members.keys()):
            for m in team_members[slug]:
                ws.append([slug, m.get("login", ""), m.get("role", "")])

    def _write_github_outside_collaborators_table(
        self, ws, outside_collaborators: list[dict]
    ) -> None:
        ws.append(["Outside Collaborators"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Login", "Name", "Email"])
        apply_header_style(ws, ws.max_row)

        for c in outside_collaborators:
            ws.append([
                c.get("login", ""),
                c.get("name", "") or "",
                c.get("email", "") or "",
            ])

    # --- Jira-specific tables ---

    def _write_jira_data_tables(self, ws, data: dict) -> None:
        self._write_jira_users_table(ws, data.get("users", []))
        ws.append([])
        self._write_jira_groups_table(
            ws, data.get("groups", []), data.get("group_members", {})
        )
        ws.append([])
        self._write_jira_group_members_table(ws, data.get("group_members", {}))

    def _write_jira_users_table(self, ws, users: list[dict]) -> None:
        ws.append(["Users"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Display Name", "Email", "Account Type", "External"])
        apply_header_style(ws, ws.max_row)

        for u in users:
            ws.append([
                u.get("display_name", ""),
                u.get("email", ""),
                u.get("account_type", ""),
                "Yes" if u.get("is_external") else "No",
            ])

    def _write_jira_groups_table(
        self, ws, groups: list[dict], group_members: dict[str, list]
    ) -> None:
        ws.append(["Groups"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Name", "Members"])
        apply_header_style(ws, ws.max_row)

        for g in groups:
            name = g.get("name", "")
            members = group_members.get(name, [])
            ws.append([name, len(members)])

    def _write_jira_group_members_table(self, ws, group_members: dict[str, list]) -> None:
        ws.append(["Group Members"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        ws.append(["Group", "Account ID", "Display Name"])
        apply_header_style(ws, ws.max_row)

        for group_name in sorted(group_members.keys()):
            for m in group_members[group_name]:
                ws.append([
                    group_name,
                    m.get("account_id", ""),
                    m.get("display_name", ""),
                ])
