"""XLSX export for the KPI Dashboard widget."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.config import ScoringConfig
from app.core.services.export_helpers import (
    DEFAULT_GREEN_THRESHOLD,
    DEFAULT_YELLOW_THRESHOLD,
    THIN_BORDER,
    apply_header_style,
    apply_indicator_traffic_light,
    apply_row_style,
    apply_score_traffic_light,
    format_month_header,
    freeze_panes,
    save_to_bytes,
    set_column_widths,
)
from app.modules.scorecard.models.global_metrics import GlobalMetricsRecord
from app.modules.scorecard.services.export_definitions import get_metric_rows

MANUAL_KPI_FIELDS = [
    "name",
    "scope",
    "responsible",
    "methodology",
    "formula",
    "target",
    "periodicity",
]
MANUAL_KPI_HEADERS = [
    "Name",
    "Scope",
    "Responsible",
    "Methodology",
    "Formula",
    "Target",
    "Periodicity",
]


ISO_CYCLE_MONTHS = 12


def generate_iso_periods(start_year: int, start_month: int) -> list[tuple[int, int]]:
    """Return exactly 12 (year, month) tuples for a fixed ISO cycle."""
    periods: list[tuple[int, int]] = []
    year, month = int(start_year), int(start_month)
    for _ in range(ISO_CYCLE_MONTHS):
        periods.append((year, month))
        month += 1
        if month > 12:
            month = 1
            year += 1
    return periods


class KpiExportService:
    """Generates XLSX with Global Scorecard and manual KPI sheets for the KPI Dashboard widget."""

    def __init__(self, config: ScoringConfig) -> None:
        self.config = config
        self._green = self._get_threshold("threshold_green", DEFAULT_GREEN_THRESHOLD)
        self._yellow = self._get_threshold("threshold_yellow", DEFAULT_YELLOW_THRESHOLD)

    def build_xlsx(
        self,
        *,
        global_by_period: dict[tuple[int, int], GlobalMetricsRecord | None],
        manual_rows: list,
        start_year: int,
        start_month: int,
    ) -> BytesIO:
        """Build workbook with two sheets: Global Scorecard and KPIs manuales."""
        periods = generate_iso_periods(start_year, start_month)

        wb = Workbook()
        ws_scorecard = wb.active
        ws_scorecard.title = "Global Scorecard"

        self._write_scorecard_sheet(ws_scorecard, periods, global_by_period)

        ws_manual = wb.create_sheet(title="KPIs manuales")
        self._write_manual_kpis_sheet(ws_manual, periods, manual_rows)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return save_to_bytes(wb)

    def _write_scorecard_sheet(
        self,
        ws,
        periods: list[tuple[int, int]],
        global_by_period: dict[tuple[int, int], GlobalMetricsRecord | None],
    ) -> None:
        """Write hierarchical scorecard table onto the sheet."""
        if periods:
            start_year = periods[0][0]
            end_year = periods[-1][0]
            ws.append([f"Global Dashboard — Ciclo {start_year}–{end_year}"])
        else:
            ws.append(["Global Dashboard"])
        ws.append([])

        metric_rows = get_metric_rows()

        header = ["Name", "Description", "Formula", "Target"]
        for year, month in periods:
            header.append(format_month_header(year, month))
        ws.append(header)
        header_row = ws.max_row
        apply_header_style(ws, header_row)

        for metric_row in metric_rows:
            level = metric_row["level"]
            indent = "  " * level
            target = self._get_target_for_metric(metric_row["key"], level)

            row_data = [
                f"{indent}{metric_row['name']}",
                metric_row["description"],
                metric_row["formula"],
                target if target else "-",
            ]

            for period in periods:
                record = global_by_period.get(period)
                row_data.append(self._extract_global_value(metric_row["key"], level, record))

            ws.append(row_data)
            current_row = ws.max_row
            apply_row_style(ws, current_row, level)

            for col_idx, _period in enumerate(periods, start=5):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = THIN_BORDER
                if level <= 1:
                    apply_score_traffic_light(cell, cell.value, self._green, self._yellow)
                else:
                    apply_indicator_traffic_light(
                        cell,
                        cell.value,
                        self._green / 100,
                        self._yellow / 100,
                    )

        freeze_panes(ws, header_row + 1, 5)

        widths = {"A": 35, "B": 50, "C": 45, "D": 12}
        for i, _ in enumerate(periods):
            widths[get_column_letter(5 + i)] = 12
        set_column_widths(ws, widths)

    def _write_manual_kpis_sheet(
        self,
        ws,
        periods: list[tuple[int, int]],
        manual_rows: list,
    ) -> None:
        """Write manual KPI rows with field columns followed by month value columns."""
        header = MANUAL_KPI_HEADERS[:]
        for year, month in periods:
            header.append(format_month_header(year, month))
        ws.append(header)
        apply_header_style(ws, ws.max_row)

        for row in manual_rows:
            data = row.data if hasattr(row, "data") else {}
            row_data = [data.get(field) for field in MANUAL_KPI_FIELDS]

            for year, month in periods:
                month_key = f"m{month:02d}"
                row_data.append(data.get(month_key))

            ws.append(row_data)
            for cell in ws[ws.max_row]:
                cell.border = THIN_BORDER

        freeze_panes(ws, 2, len(MANUAL_KPI_FIELDS) + 1)

        widths = {}
        for i in range(len(MANUAL_KPI_FIELDS)):
            widths[get_column_letter(i + 1)] = 20
        for i, _ in enumerate(periods):
            widths[get_column_letter(len(MANUAL_KPI_FIELDS) + 1 + i)] = 12
        set_column_widths(ws, widths)

    @staticmethod
    def _extract_global_value(
        key: str, level: int, record: GlobalMetricsRecord | None
    ) -> float | None:
        """Extract averaged value from a GlobalMetricsRecord."""
        if record is None:
            return None
        if level == 0:
            val = record.scores.score.value
            return round(val, 1) if val is not None else None
        if level == 1:
            score_val = getattr(record.scores, key, None)
            if score_val and score_val.value is not None:
                return round(score_val.value, 1)
            return None
        indicator_val = getattr(record.indicators, key, None)
        if indicator_val and indicator_val.value is not None:
            return round(indicator_val.value, 1)
        return None

    def _get_target_for_metric(self, key: str, level: int) -> str | None:
        """Get display-friendly target string for a metric row."""
        if level <= 1:
            return str(int(self._green))
        try:
            return str(self.config.get_target(key))
        except (KeyError, ValueError):
            return None

    def _get_threshold(self, name: str, default: float) -> float:
        """Get threshold constant from config, falling back to default."""
        try:
            val = self.config.get_constant(name)
            return val if val > 0 else default
        except (KeyError, ValueError):
            return default
