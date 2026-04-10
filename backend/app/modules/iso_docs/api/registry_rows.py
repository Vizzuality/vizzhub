"""Registry row CRUD + reorder + export endpoints."""

from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from io import BytesIO, StringIO
from typing import Annotated
from uuid import UUID

import httpx
import openpyxl
import structlog
from fastapi import APIRouter, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import delete as sql_delete, select

from app.core.api.deps import CurrentUser, DBSession
from app.modules.iso_docs.api.deps import IsoDocsEditor, check_user_access
from app.modules.iso_docs.models.metadata import IsoDocMetadataDB
from app.modules.iso_docs.models.node import IsoDocNodeDB
from app.modules.iso_docs.models.registry_row import RegistryRowDB
from app.modules.iso_docs.models.registry_type import RegistryTypeDB

_REGISTRY_TYPE_NOT_FOUND = "Registry type not found"
from app.modules.iso_docs.schemas.registry import (
    RegistryRowCreate,
    RegistryRowReorder,
    RegistryRowResponse,
    RegistryRowUpdate,
)
from app.modules.iso_docs.services.drive_export_service import (
    DRIVE_API,
    DRIVE_TIMEOUT,
    DRIVE_UPLOAD_API,
)
from app.modules.iso_docs.services.registry_attachment_service import (
    get_attachment_url,
)
from app.modules.iso_docs.services.registry_service import (
    compute_row_fields,
    get_next_row_index,
    strip_computed_keys,
    validate_row_data,
)

logger = structlog.get_logger()

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

router = APIRouter()


async def _get_registry_node(db, node_id: UUID) -> IsoDocNodeDB:
    result = await db.execute(
        select(IsoDocNodeDB).where(
            IsoDocNodeDB.id == node_id, IsoDocNodeDB.type.in_(["registry", "widget"])
        )
    )
    node = result.scalar_one_or_none()
    if not node:
        raise HTTPException(status_code=404, detail="Registry node not found")
    return node


async def _get_registry_type(db, type_id: UUID | None) -> RegistryTypeDB | None:
    if type_id is None:
        return None
    result = await db.execute(
        select(RegistryTypeDB).where(RegistryTypeDB.id == type_id)
    )
    rt = result.scalar_one_or_none()
    if not rt:
        raise HTTPException(status_code=404, detail=_REGISTRY_TYPE_NOT_FOUND)
    return rt


async def _fetch_rows(
    db, node_id: UUID, year: int | None = None, sort_key: str | None = None
) -> list:
    """Fetch ordered rows for a registry node, optionally filtered by year."""
    query = select(RegistryRowDB).where(RegistryRowDB.node_id == node_id)
    if year is not None:
        query = query.where(RegistryRowDB.year == year)
    if sort_key:
        query = query.order_by(
            RegistryRowDB.data[sort_key].astext, RegistryRowDB.row_index
        )
    else:
        query = query.order_by(RegistryRowDB.row_index)
    result = await db.execute(query)
    return list(result.scalars())


async def _fetch_metadata(db, node_id: UUID) -> IsoDocMetadataDB | None:
    result = await db.execute(
        select(IsoDocMetadataDB).where(IsoDocMetadataDB.node_id == node_id)
    )
    return result.scalar_one_or_none()


def _group_rows_by_year(rows: list) -> dict[int, list]:
    grouped: dict[int, list] = {}
    for r in rows:
        grouped.setdefault(r.year, []).append(r)
    return grouped


@router.get(
    "/registries/{node_id}/years",
    responses={404: {"description": "Registry node not found"}},
)
async def list_years(
    node_id: UUID, db: DBSession, user: CurrentUser
) -> list[int]:
    await check_user_access(db, node_id, user)
    node = await _get_registry_node(db, node_id)
    result = await db.execute(
        select(RegistryRowDB.year)
        .where(RegistryRowDB.node_id == node.id, RegistryRowDB.year.is_not(None))
        .distinct()
        .order_by(RegistryRowDB.year.desc())
    )
    return list(result.scalars())


def _enrich_row(
    row, schema: list[dict]
) -> RegistryRowResponse:
    """Serialize a row with computed field values and attachment URLs injected."""
    resp = RegistryRowResponse.model_validate(row)
    resp.data = compute_row_fields(schema, resp.data)
    for att in resp.attachments:
        att.url = get_attachment_url(att.s3_key)
    return resp


def _enrich_rows(
    rows: list, schema: list[dict]
) -> list[RegistryRowResponse]:
    """Serialize rows with computed field values and attachment URLs injected."""
    return [_enrich_row(r, schema) for r in rows]


@router.get(
    "/registries/{node_id}/rows",
    responses={404: {"description": "Registry node not found"}},
)
async def list_rows(
    node_id: UUID,
    db: DBSession,
    user: CurrentUser,
    year: Annotated[int | None, Query()] = None,
) -> list[RegistryRowResponse]:
    await check_user_access(db, node_id, user)
    node = await _get_registry_node(db, node_id)
    rt = await _get_registry_type(db, node.registry_type_id)
    sort_key = rt.default_sort_key if rt else None
    rows = await _fetch_rows(db, node.id, year, sort_key)
    return _enrich_rows(rows, rt.schema if rt else [])


@router.post(
    "/registries/{node_id}/rows",
    status_code=201,
    responses={
        404: {"description": "Registry node or type not found"},
        400: {"description": "Year is required for yearly registries"},
        422: {"description": "Row data validation failed"},
    },
)
async def create_row(
    node_id: UUID,
    data: RegistryRowCreate,
    db: DBSession,
    user: IsoDocsEditor,
) -> RegistryRowResponse:
    node = await _get_registry_node(db, node_id)
    rt = await _get_registry_type(db, node.registry_type_id)
    schema = rt.schema if rt else []

    if rt and rt.is_yearly and data.year is None:
        raise HTTPException(status_code=400, detail="Year is required for yearly registries")

    clean_data = strip_computed_keys(schema, data.data) if schema else data.data
    if schema:
        errors = validate_row_data(schema, clean_data)
        if errors:
            raise HTTPException(status_code=422, detail=errors)

    row_index = await get_next_row_index(db, node.id, data.year)
    row = RegistryRowDB(
        node_id=node.id,
        year=data.year,
        row_index=row_index,
        data=clean_data,
        created_by_id=UUID(user.user_id),
        updated_by_id=UUID(user.user_id),
    )
    db.add(row)
    await db.flush()
    await db.refresh(row)
    logger.info("registry_row_created", node_id=str(node_id), row_id=str(row.id))
    return _enrich_row(row, schema)


@router.patch(
    "/registries/{node_id}/rows/{row_id}",
    responses={
        404: {"description": "Row not found"},
        422: {"description": "Row data validation failed"},
    },
)
async def update_row(
    node_id: UUID,
    row_id: UUID,
    data: RegistryRowUpdate,
    db: DBSession,
    user: IsoDocsEditor,
) -> RegistryRowResponse:
    node = await _get_registry_node(db, node_id)
    rt = await _get_registry_type(db, node.registry_type_id)
    schema = rt.schema if rt else []

    result = await db.execute(
        select(RegistryRowDB).where(
            RegistryRowDB.id == row_id, RegistryRowDB.node_id == node.id
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Row not found")

    clean_update = strip_computed_keys(schema, data.data) if schema else data.data
    merged = {**row.data, **clean_update}
    if schema:
        errors = validate_row_data(schema, merged)
        if errors:
            raise HTTPException(status_code=422, detail=errors)

    row.data = merged
    row.updated_by_id = UUID(user.user_id)
    await db.flush()
    await db.refresh(row)
    logger.info("registry_row_updated", node_id=str(node_id), row_id=str(row_id))
    return _enrich_row(row, schema)


@router.delete(
    "/registries/{node_id}/rows/{row_id}",
    responses={404: {"description": "Row not found"}},
)
async def delete_row(
    node_id: UUID,
    row_id: UUID,
    db: DBSession,
    user: IsoDocsEditor,
) -> dict:
    await _get_registry_node(db, node_id)

    result = await db.execute(
        select(RegistryRowDB).where(
            RegistryRowDB.id == row_id, RegistryRowDB.node_id == node_id
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Row not found")

    await db.delete(row)
    await db.flush()
    logger.info("registry_row_deleted", node_id=str(node_id), row_id=str(row_id))
    return {"ok": True}


@router.put(
    "/registries/{node_id}/rows/reorder",
    responses={404: {"description": "Registry node or row not found"}},
)
async def reorder_rows(
    node_id: UUID,
    data: RegistryRowReorder,
    db: DBSession,
    user: IsoDocsEditor,
) -> dict:
    await _get_registry_node(db, node_id)

    result = await db.execute(
        select(RegistryRowDB).where(
            RegistryRowDB.id.in_(data.row_ids),
            RegistryRowDB.node_id == node_id,
        )
    )
    rows_by_id = {r.id: r for r in result.scalars()}

    for idx, row_id in enumerate(data.row_ids):
        row = rows_by_id.get(row_id)
        if not row:
            raise HTTPException(status_code=404, detail=f"Row {row_id} not found")
        row.row_index = idx

    await db.flush()
    logger.info("registry_rows_reordered", node_id=str(node_id), count=len(data.row_ids))
    return {"ok": True}


@router.get(
    "/registries/{node_id}/export",
    responses={404: {"description": "Registry node or type not found"}},
)
async def export_registry(
    node_id: UUID,
    db: DBSession,
    user: CurrentUser,
    year: Annotated[int | None, Query()] = None,
    format: Annotated[str, Query()] = "xlsx",
) -> StreamingResponse:
    await check_user_access(db, node_id, user)
    node = await _get_registry_node(db, node_id)
    rt = await _get_registry_type(db, node.registry_type_id)
    if rt is None:
        raise HTTPException(status_code=404, detail=_REGISTRY_TYPE_NOT_FOUND)
    rows = await _fetch_rows(db, node.id, year)
    metadata = await _fetch_metadata(db, node.id)

    columns = rt.schema
    headers = [col["label"] for col in columns]
    base_name = f"{node.title} ({year})" if year else node.title

    if format == "csv":
        buf = StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        for row in rows:
            enriched = compute_row_fields(columns, row.data)
            writer.writerow([enriched.get(col["key"], "") for col in columns])

        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{base_name}.csv"'},
        )

    if rt.is_yearly and year is None:
        all_rows = await _fetch_rows(db, node.id)
        xlsx_buf = _build_xlsx_multiyear(
            columns, _group_rows_by_year(all_rows), metadata,
        )
    else:
        xlsx_buf = _build_xlsx(node.title, columns, rows, metadata)
    return StreamingResponse(
        xlsx_buf,
        media_type=XLSX_CONTENT_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{base_name}.xlsx"'},
    )


def _coerce_csv_value(raw: str, col_type: str) -> object:
    """Convert a CSV string to the appropriate Python type."""
    if raw == "":
        return None
    if col_type == "number":
        try:
            return int(raw)
        except ValueError:
            try:
                return float(raw)
            except ValueError:
                return raw
    if col_type == "boolean":
        return raw.lower() in ("true", "yes", "1")
    return raw


def _parse_csv(text: str, columns: list[dict]) -> list[dict]:
    """Parse CSV text into validated row dicts. Raises HTTPException on errors."""
    computed_labels = {col["label"] for col in columns if col["type"] == "computed"}
    label_to_col = {col["label"]: col for col in columns if col["type"] != "computed"}
    reader = csv.DictReader(StringIO(text))

    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV has no headers")

    unknown = [
        h for h in reader.fieldnames
        if h and h not in label_to_col and h not in computed_labels
    ]
    if unknown:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown columns: {', '.join(repr(h) for h in unknown)}. Expected: {', '.join(label_to_col.keys())}",
        )

    parsed_rows: list[dict] = []
    for line_num, csv_row in enumerate(reader, start=2):
        data: dict = {}
        for label, raw in csv_row.items():
            col = label_to_col.get(label)
            if not col:
                continue
            data[col["key"]] = _coerce_csv_value(raw or "", col["type"])

        errors = validate_row_data(columns, data)
        if errors:
            raise HTTPException(
                status_code=400,
                detail=f"Row {line_num}: {'; '.join(errors)}",
            )
        parsed_rows.append(data)
    return parsed_rows


@router.post(
    "/registries/{node_id}/import",
    responses={
        404: {"description": "Registry node or type not found"},
        400: {"description": "Invalid CSV file or data"},
    },
)
async def import_registry(
    node_id: UUID,
    file: UploadFile,
    db: DBSession,
    user: IsoDocsEditor,
    year: Annotated[int | None, Query()] = None,
) -> dict:
    if not file.filename or not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files are accepted")

    node = await _get_registry_node(db, node_id)
    rt = await _get_registry_type(db, node.registry_type_id)
    if rt is None:
        raise HTTPException(status_code=404, detail=_REGISTRY_TYPE_NOT_FOUND)

    content = await file.read()
    text = content.decode("utf-8-sig")
    parsed_rows = _parse_csv(text, rt.schema)

    await db.execute(
        select(RegistryRowDB)
        .where(RegistryRowDB.node_id == node.id)
        .with_for_update()
    )

    if year is not None:
        await db.execute(
            sql_delete(RegistryRowDB).where(
                RegistryRowDB.node_id == node.id,
                RegistryRowDB.year == year,
            )
        )
    else:
        await db.execute(
            sql_delete(RegistryRowDB).where(RegistryRowDB.node_id == node.id)
        )

    for idx, data in enumerate(parsed_rows):
        db.add(
            RegistryRowDB(
                node_id=node.id,
                year=year,
                row_index=idx,
                data=data,
                created_by_id=UUID(user.user_id),
                updated_by_id=UUID(user.user_id),
            )
        )

    await db.flush()
    logger.info(
        "registry_csv_imported",
        node_id=str(node_id),
        rows_imported=len(parsed_rows),
        year=year,
    )
    return {"imported": len(parsed_rows)}


@router.post(
    "/registries/{node_id}/copy-year",
    responses={
        404: {"description": "Registry node or type not found"},
        400: {"description": "Invalid copy request"},
    },
)
async def copy_year(
    node_id: UUID,
    db: DBSession,
    user: IsoDocsEditor,
    source_year: Annotated[int, Query()],
    target_year: Annotated[int, Query()],
) -> dict:
    node = await _get_registry_node(db, node_id)
    rt = await _get_registry_type(db, node.registry_type_id)

    if rt and not rt.is_yearly:
        raise HTTPException(status_code=400, detail="Only yearly registries support copy")
    if source_year == target_year:
        raise HTTPException(status_code=400, detail="Source and target year must differ")

    existing = await db.execute(
        select(RegistryRowDB).where(
            RegistryRowDB.node_id == node.id, RegistryRowDB.year == target_year
        ).limit(1)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Target year already has data")

    source_rows = await _fetch_rows(db, node.id, source_year)
    if not source_rows:
        raise HTTPException(status_code=400, detail="Source year has no data to copy")

    user_id = UUID(user.user_id)
    for idx, src in enumerate(source_rows):
        db.add(RegistryRowDB(
            node_id=node.id,
            year=target_year,
            row_index=idx,
            data=dict(src.data),
            created_by_id=user_id,
            updated_by_id=user_id,
        ))

    await db.flush()
    logger.info(
        "registry_year_copied",
        node_id=str(node_id),
        source_year=source_year,
        target_year=target_year,
        rows_copied=len(source_rows),
    )
    return {"copied": len(source_rows)}


HEADER_FONT = Font(name="Calibri", size=10, bold=True)
HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
META_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="555555")
META_VALUE_FONT = Font(name="Calibri", size=10)

STATUS_LABELS = {"draft": "Draft", "approved": "Approved", "under_review": "Under review"}
CATEGORY_LABELS = {
    "manual": "Manual", "policy": "Policy", "procedure": "Procedure",
    "plan": "Plan", "record": "Record", "report": "Report",
}
CLASSIFICATION_LABELS = {
    "internal_use": "Internal use", "confidential": "Confidential",
}


def _build_meta_rows(metadata) -> list[tuple[str, str]]:
    """Extract displayable label-value pairs from metadata."""
    if metadata is None:
        return []
    pairs: list[tuple[str, str]] = []
    if metadata.code:
        pairs.append(("Code", metadata.code))
    if metadata.category:
        pairs.append(("Category", CATEGORY_LABELS.get(metadata.category, metadata.category)))
    if metadata.status:
        pairs.append(("Status", STATUS_LABELS.get(metadata.status, metadata.status)))
    if metadata.classification:
        pairs.append(("Classification", CLASSIFICATION_LABELS.get(
            metadata.classification, metadata.classification,
        )))
    if metadata.standard:
        pairs.append(("Standard", ", ".join(metadata.standard)))
    if metadata.clauses:
        pairs.append(("Clauses", ", ".join(metadata.clauses)))
    if metadata.doc_version:
        pairs.append(("Version", metadata.doc_version))
    if metadata.document_date:
        pairs.append(("Date", str(metadata.document_date)))
    return pairs


def _populate_sheet(
    ws, columns: list[dict], rows: list, metadata=None,
) -> None:
    """Write optional metadata header, column headers + rows with styling."""
    meta_rows = _build_meta_rows(metadata)
    for label, value in meta_rows:
        ws.append([label, value])
    if meta_rows:
        ws.append([])

    header_row_num = len(meta_rows) + 2 if meta_rows else 1
    headers = [col["label"] for col in columns]
    ws.append(headers)
    for row in rows:
        enriched = compute_row_fields(columns, row.data)
        ws.append([enriched.get(col["key"], "") for col in columns])

    for row_num in range(1, len(meta_rows) + 1):
        ws.cell(row=row_num, column=1).font = META_LABEL_FONT
        ws.cell(row=row_num, column=2).font = META_VALUE_FONT

    for cell in ws[header_row_num]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    for row in ws.iter_rows(min_row=header_row_num + 1):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGNMENT

    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in ws.iter_rows(
            min_row=header_row_num + 1, min_col=col_idx, max_col=col_idx,
        ):
            val = row[0].value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    ws.freeze_panes = f"A{header_row_num + 1}"


def _build_xlsx(
    node_title: str, columns: list[dict], rows: list, metadata=None,
) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = node_title[:31]
    _populate_sheet(ws, columns, rows, metadata)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_xlsx_multiyear(
    columns: list[dict], rows_by_year: dict[int, list],
    metadata=None,
) -> BytesIO:
    """Build an XLSX with one tab per year, most recent first."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for year in sorted(rows_by_year, reverse=True):
        ws = wb.create_sheet(title=str(year))
        _populate_sheet(ws, columns, rows_by_year[year], metadata)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def _resolve_drive_parent(
    db,
    node: IsoDocNodeDB,
    http: httpx.AsyncClient,
    auth_header: dict[str, str],
) -> str:
    """Walk up ancestors to find a Drive folder, creating missing ones."""
    from app.modules.iso_docs.models.drive_mapping import IsoDocDriveMappingDB
    from app.core.services.integration_token_service import IntegrationTokenService
    from app.modules.iso_docs.services.google_drive_oauth import PROVIDER

    ancestors: list[IsoDocNodeDB] = []
    current = node
    while current.parent_id:
        result = await db.execute(
            select(IsoDocNodeDB).where(IsoDocNodeDB.id == current.parent_id)
        )
        parent = result.scalar_one_or_none()
        if not parent:
            break
        mapping_result = await db.execute(
            select(IsoDocDriveMappingDB).where(
                IsoDocDriveMappingDB.node_id == parent.id
            )
        )
        mapping = mapping_result.scalar_one_or_none()
        if mapping:
            drive_parent_id = mapping.drive_file_id
            break
        ancestors.append(parent)
        current = parent
    else:
        root_folder_id = await IntegrationTokenService.get_setting(
            db, PROVIDER, "root_folder_id"
        )
        if not root_folder_id:
            raise HTTPException(
                status_code=400,
                detail="Google Drive root folder not configured",
            )
        drive_parent_id = root_folder_id

    now = datetime.now(timezone.utc)
    for ancestor in reversed(ancestors):
        body = {
            "name": ancestor.title,
            "mimeType": "application/vnd.google-apps.folder",
            "parents": [drive_parent_id],
        }
        resp = await http.post(
            DRIVE_API,
            json=body,
            headers=auth_header,
            params={"fields": "id", "supportsAllDrives": "true"},
        )
        resp.raise_for_status()
        new_folder_id = resp.json()["id"]
        db.add(
            IsoDocDriveMappingDB(
                node_id=ancestor.id,
                drive_file_id=new_folder_id,
                drive_file_type="folder",
                last_exported_at=now,
            )
        )
        await db.flush()
        drive_parent_id = new_folder_id

    return drive_parent_id


@router.post(
    "/registries/{node_id}/export-drive",
    responses={
        404: {"description": "Registry node or type not found"},
        400: {"description": "Google Drive not connected"},
    },
)
async def export_registry_to_drive(
    node_id: UUID,
    db: DBSession,
    user: IsoDocsEditor,
    year: Annotated[int | None, Query()] = None,
) -> dict:
    from app.modules.iso_docs.models.drive_mapping import IsoDocDriveMappingDB
    from app.modules.iso_docs.services.google_drive_oauth import GoogleDriveOAuth

    access_token = await GoogleDriveOAuth.get_valid_token(db)
    if not access_token:
        raise HTTPException(status_code=400, detail="Google Drive not connected")

    node = await _get_registry_node(db, node_id)
    rt = await _get_registry_type(db, node.registry_type_id)
    doc_metadata = await _fetch_metadata(db, node.id)

    file_name = f"{node.title} ({year})" if year else node.title
    if rt.is_yearly and year is None:
        all_rows = await _fetch_rows(db, node.id)
        xlsx_buf = _build_xlsx_multiyear(
            rt.schema, _group_rows_by_year(all_rows), doc_metadata,
        )
        row_count = len(all_rows)
    else:
        rows = await _fetch_rows(db, node.id, year)
        xlsx_buf = _build_xlsx(node.title, rt.schema, rows, doc_metadata)
        row_count = len(rows)

    async with httpx.AsyncClient(timeout=DRIVE_TIMEOUT) as http:
        auth_header = {"Authorization": f"Bearer {access_token}"}

        parent_drive_id = await _resolve_drive_parent(db, node, http, auth_header)

        existing_mapping = await db.execute(
            select(IsoDocDriveMappingDB).where(
                IsoDocDriveMappingDB.node_id == node.id
            )
        )
        existing = existing_mapping.scalar_one_or_none()
        existing_drive_id = existing.drive_file_id if existing else None

        if existing_drive_id:
            check = await http.get(
                f"{DRIVE_API}/{existing_drive_id}",
                headers=auth_header,
                params={"fields": "id,trashed", "supportsAllDrives": "true"},
            )
            if check.status_code == 200 and not check.json().get("trashed"):
                await http.patch(
                    f"{DRIVE_API}/{existing_drive_id}",
                    json={"name": file_name},
                    headers=auth_header,
                    params={"supportsAllDrives": "true"},
                )
                await http.patch(
                    f"{DRIVE_UPLOAD_API}/{existing_drive_id}?uploadType=media",
                    content=xlsx_buf.read(),
                    headers={
                        **auth_header,
                        "Content-Type": XLSX_CONTENT_TYPE,
                    },
                    params={"supportsAllDrives": "true"},
                )
                drive_file_id = existing_drive_id
            else:
                existing_drive_id = None

        if not existing_drive_id:
            drive_meta = {
                "name": file_name,
                "parents": [parent_drive_id],
                "mimeType": "application/vnd.google-apps.spreadsheet",
            }
            resp = await http.post(
                f"{DRIVE_UPLOAD_API}?uploadType=multipart",
                headers=auth_header,
                files={
                    "metadata": (
                        None, json.dumps(drive_meta).encode(), "application/json",
                    ),
                    "file": (None, xlsx_buf.read(), XLSX_CONTENT_TYPE),
                },
                params={"fields": "id", "supportsAllDrives": "true"},
            )
            resp.raise_for_status()
            drive_file_id = resp.json()["id"]

    now = datetime.now(timezone.utc)
    if existing:
        existing.drive_file_id = drive_file_id
        existing.last_exported_at = now
    else:
        db.add(
            IsoDocDriveMappingDB(
                node_id=node.id,
                drive_file_id=drive_file_id,
                drive_file_type="spreadsheet",
                last_exported_at=now,
            )
        )
    await db.flush()

    logger.info(
        "registry_exported_to_drive",
        node_id=str(node_id),
        drive_file_id=drive_file_id,
        rows=row_count,
    )
    return {"drive_file_id": drive_file_id}
