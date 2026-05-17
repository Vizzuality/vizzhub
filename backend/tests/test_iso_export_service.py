"""Tests for ISO export service — XLSX generation."""

from datetime import UTC, datetime
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook

from app.modules.iso.services.export_service import IsoExportService


def _make_snapshot(
    captured_at=None,
    domain="test.com",
    users=None,
    groups=None,
    group_members=None,
    role_assignments=None,
):
    """Build a fake snapshot dict matching AccessSnapshotDB shape."""
    return {
        "id": str(uuid4()),
        "provider": "google_workspace",
        "captured_at": captured_at or datetime(2026, 6, 15, 10, 0, tzinfo=UTC),
        "data_version": "1",
        "source_metadata": {"domain": domain},
        "data": {
            "users": users
            or [
                {
                    "id": "u1",
                    "name": "Alice",
                    "email": "alice@test.com",
                    "suspended": False,
                    "org_unit_path": "/",
                },
            ],
            "groups": groups
            or [
                {"id": "g1", "name": "Engineering", "email": "eng@test.com"},
            ],
            "group_members": group_members
            or {
                "eng@test.com": [
                    {"email": "alice@test.com", "role": "MEMBER", "type": "USER"},
                ],
            },
            "role_assignments": role_assignments
            or [
                {
                    "role_id": "r1",
                    "user_id": "u1",
                    "role_name": "Super Admin",
                    "user_email": "alice@test.com",
                },
            ],
        },
        "summary": {
            "total_users": 1,
            "total_admins": 1,
            "total_groups": 1,
            "external_members": 0,
        },
    }


def _make_review(
    status="signed",
    notes="All good",
    reviewer_email=None,
    signed_by_email=None,
    signed_at=None,
    diff_summary=None,
):
    return {
        "id": str(uuid4()),
        "status": status,
        "scope": "All users and groups",
        "notes": notes,
        "reviewer_email": reviewer_email or "admin@test.com",
        "signed_by_email": signed_by_email or "admin@test.com",
        "signed_at": signed_at or datetime(2026, 6, 16, 12, 0, tzinfo=UTC),
        "diff_summary": diff_summary
        or {
            "total_changes": 1,
            "new_user": 1,
            "removed_user": 0,
            "role_change": 0,
            "new_external": 0,
            "group_membership_change": 0,
        },
    }


def _make_action(action_taken="accepted", justification="Approved"):
    return {
        "subject_label": "Alice",
        "subject_type": "user",
        "change_type": "new_user",
        "previous_value": None,
        "current_value": {"email": "alice@test.com"},
        "action_taken": action_taken,
        "justification": justification,
        "exception_until": None,
    }


class TestIsoExportService:
    def test_generates_workbook_with_snapshot_tab(self):
        service = IsoExportService()
        snapshot = _make_snapshot()
        review = _make_review()
        actions = [_make_action()]

        output = service.export_snapshots(snapshots_with_reviews=[(snapshot, review, actions)])

        assert isinstance(output, BytesIO)
        wb = load_workbook(output)
        assert "Summary" not in wb.sheetnames
        assert len(wb.sheetnames) == 1

    def test_snapshot_tab_has_iso_header(self):
        service = IsoExportService()
        snapshot = _make_snapshot(domain="acme.com")
        review = _make_review()

        output = service.export_snapshots(snapshots_with_reviews=[(snapshot, review, [])])
        wb = load_workbook(output)
        ws = wb[wb.sheetnames[0]]

        labels = [ws.cell(row=r, column=1).value for r in range(1, 16)]
        assert "Organization" in labels
        assert "Provider" in labels
        assert "Total Users" in labels
        assert "Total Admins" in labels
        assert "Review Scope" in labels

        org_row = labels.index("Organization") + 1
        assert ws.cell(row=org_row, column=2).value == "acme.com"

    def test_snapshot_tab_has_users_table(self):
        service = IsoExportService()
        snapshot = _make_snapshot()
        review = _make_review()

        output = service.export_snapshots(snapshots_with_reviews=[(snapshot, review, [])])
        wb = load_workbook(output)
        ws = wb[wb.sheetnames[0]]

        all_values = [
            ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        assert "alice@test.com" in all_values

    def test_snapshot_tab_has_actions_table(self):
        service = IsoExportService()
        snapshot = _make_snapshot()
        review = _make_review()
        actions = [_make_action()]

        output = service.export_snapshots(snapshots_with_reviews=[(snapshot, review, actions)])
        wb = load_workbook(output)
        ws = wb[wb.sheetnames[0]]

        all_values = [
            ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        assert "Approved" in all_values

    def test_multiple_snapshots_generate_multiple_tabs(self):
        service = IsoExportService()
        snap1 = _make_snapshot(
            captured_at=datetime(2026, 3, 1, tzinfo=UTC),
        )
        snap2 = _make_snapshot(
            captured_at=datetime(2026, 6, 1, tzinfo=UTC),
        )
        review1 = _make_review()
        review2 = _make_review()

        output = service.export_snapshots(
            snapshots_with_reviews=[
                (snap1, review1, []),
                (snap2, review2, []),
            ]
        )
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 2

    def test_snapshot_without_review(self):
        service = IsoExportService()
        snapshot = _make_snapshot()

        output = service.export_snapshots(snapshots_with_reviews=[(snapshot, None, [])])
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 1
