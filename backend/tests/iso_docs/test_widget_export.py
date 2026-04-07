"""Tests for KpiExportService XLSX generation."""

from __future__ import annotations

from unittest.mock import MagicMock

import pytest
from openpyxl import load_workbook

from app.modules.iso_docs.services.kpi_export_service import KpiExportService


@pytest.fixture
def mock_config():
    config = MagicMock()
    config.get_target = MagicMock(side_effect=lambda k: {"spi": 0.8, "cpi": 0.8}.get(k, 0))
    config.get_constant = MagicMock(
        side_effect=lambda k: {"threshold_green": 80, "threshold_yellow": 60}.get(k, 0)
    )
    config.global_weights = MagicMock()
    config.global_weights.time = 0.12
    config.global_weights.cost = 0.12
    config.global_weights.quality = 0.15
    config.global_weights.value = 0.10
    config.global_weights.satisfaction = 0.15
    config.global_weights.flow = 0.12
    config.global_weights.engineering = 0.12
    config.global_weights.risk = 0.12
    return config


def _build_service(mock_config) -> KpiExportService:
    return KpiExportService(mock_config)


def test_generates_xlsx_with_two_sheets(mock_config):
    """Build with no data and verify both sheet names exist."""
    service = _build_service(mock_config)

    result = service.build_xlsx(
        global_by_period={},
        manual_rows=[],
        start_year=2025,
        start_month=3,
        end_year=2026,
        end_month=2,
    )

    wb = load_workbook(result)
    assert "Global Scorecard" in wb.sheetnames
    assert "KPIs manuales" in wb.sheetnames


def test_scorecard_sheet_has_metric_rows(mock_config):
    """Verify FINAL SCORE and P_time appear in the first column of Global Scorecard."""
    service = _build_service(mock_config)

    result = service.build_xlsx(
        global_by_period={},
        manual_rows=[],
        start_year=2025,
        start_month=3,
        end_year=2026,
        end_month=2,
    )

    wb = load_workbook(result)
    ws = wb["Global Scorecard"]

    col_a_values = [str(cell.value or "").strip() for cell in ws["A"]]

    assert any("FINAL SCORE" in v for v in col_a_values)
    assert any("P_time" in v for v in col_a_values)


def test_manual_kpis_sheet_with_rows(mock_config):
    """Verify a manual KPI row's field values and month data appear in the sheet."""
    service = _build_service(mock_config)

    mock_row = MagicMock()
    mock_row.data = {
        "name": "% formación seguridad",
        "scope": "Concienciación",
        "responsible": "RRHH",
        "methodology": "Porcentaje formados",
        "formula": "formados / total",
        "target": 0.8,
        "periodicity": "Anual",
        "m03": 0.75,
        "m04": None,
    }

    result = service.build_xlsx(
        global_by_period={},
        manual_rows=[mock_row],
        start_year=2025,
        start_month=3,
        end_year=2025,
        end_month=4,
    )

    wb = load_workbook(result)
    ws = wb["KPIs manuales"]

    all_values = [[cell.value for cell in row] for row in ws.iter_rows()]
    flat_values = [v for row in all_values for v in row]

    assert "% formación seguridad" in flat_values
    assert "RRHH" in flat_values
    assert 0.75 in flat_values


def test_month_columns_follow_iso_cycle(mock_config):
    """Verify Mar 2025 and Feb 2026 appear as column headers for a full ISO cycle."""
    service = _build_service(mock_config)

    result = service.build_xlsx(
        global_by_period={},
        manual_rows=[],
        start_year=2025,
        start_month=3,
        end_year=2026,
        end_month=2,
    )

    wb = load_workbook(result)
    ws = wb["Global Scorecard"]

    all_values = [[cell.value for cell in row] for row in ws.iter_rows()]
    flat_values = [str(v) for row in all_values for v in row if v is not None]

    assert "Mar 2025" in flat_values
    assert "Feb 2026" in flat_values
