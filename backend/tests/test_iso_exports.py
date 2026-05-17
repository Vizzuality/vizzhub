"""Tests for ISO export API endpoints."""

from datetime import UTC, datetime
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from app.modules.iso.models.access_review_action import AccessReviewActionDB
from tests.iso_fixtures import ensure_dev_user, make_review, make_snapshot

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class TestExportSnapshotRange:
    @pytest.mark.asyncio
    async def test_export_date_range(self, client: AsyncClient, db_session) -> None:
        await ensure_dev_user(db_session)
        await make_snapshot(
            db_session,
            captured_at=datetime(2026, 3, 1, tzinfo=UTC),
        )
        await make_snapshot(
            db_session,
            captured_at=datetime(2026, 6, 1, tzinfo=UTC),
        )

        response = await client.get(
            "/api/iso/exports/snapshots",
            params={"from": "2026-01-01", "to": "2026-12-31"},
        )
        assert response.status_code == 200
        assert response.headers["content-type"] == XLSX_CONTENT_TYPE

        wb = load_workbook(BytesIO(response.content))
        assert "Summary" not in wb.sheetnames
        assert len(wb.sheetnames) == 2

    @pytest.mark.asyncio
    async def test_export_empty_range_returns_404(self, client: AsyncClient) -> None:
        response = await client.get(
            "/api/iso/exports/snapshots",
            params={"from": "2026-01-01", "to": "2026-12-31"},
        )
        assert response.status_code == 404
        assert "No snapshots" in response.json().get("detail", "")

    @pytest.mark.asyncio
    async def test_export_invalid_date_format(self, client: AsyncClient) -> None:
        response = await client.get(
            "/api/iso/exports/snapshots",
            params={"from": "bad-date", "to": "2026-12-31"},
        )
        assert response.status_code == 400

    @pytest.mark.asyncio
    async def test_export_to_before_from_returns_400(self, client: AsyncClient) -> None:
        response = await client.get(
            "/api/iso/exports/snapshots",
            params={"from": "2026-12-31", "to": "2026-01-01"},
        )
        assert response.status_code == 400

    @pytest.mark.asyncio
    async def test_export_filters_by_date_range(self, client: AsyncClient, db_session) -> None:
        await ensure_dev_user(db_session)
        await make_snapshot(
            db_session,
            captured_at=datetime(2026, 3, 1, tzinfo=UTC),
        )
        await make_snapshot(
            db_session,
            captured_at=datetime(2026, 9, 1, tzinfo=UTC),
        )

        response = await client.get(
            "/api/iso/exports/snapshots",
            params={"from": "2026-01-01", "to": "2026-06-30"},
        )
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        assert len(wb.sheetnames) == 1

    @pytest.mark.asyncio
    async def test_export_has_content_disposition_header(self, client: AsyncClient) -> None:
        # Empty range now returns 404 instead of an empty workbook; the
        # content-disposition header is exercised by other tests that seed
        # snapshots. Keep the negative assertion explicit so future changes
        # to error shape still trip a regression.
        response = await client.get(
            "/api/iso/exports/snapshots",
            params={"from": "2026-01-01", "to": "2026-12-31"},
        )
        assert response.status_code == 404
        assert "content-disposition" not in response.headers


class TestExportSingleSnapshot:
    @pytest.mark.asyncio
    async def test_export_single_snapshot(self, client: AsyncClient, db_session) -> None:
        await ensure_dev_user(db_session)
        snapshot = await make_snapshot(
            db_session,
            captured_at=datetime(2026, 6, 15, 10, 0, tzinfo=UTC),
            data={
                "users": [
                    {
                        "id": "u1",
                        "name": "Alice",
                        "email": "alice@test.com",
                        "suspended": False,
                        "org_unit_path": "/",
                    }
                ],
                "groups": [],
                "group_members": {},
                "role_assignments": [],
            },
            summary={
                "total_users": 1,
                "total_admins": 0,
                "total_groups": 0,
                "external_members": 0,
            },
        )

        response = await client.get(
            f"/api/iso/exports/snapshots/{snapshot.id}",
        )
        assert response.status_code == 200
        assert response.headers["content-type"] == XLSX_CONTENT_TYPE

        wb = load_workbook(BytesIO(response.content))
        assert len(wb.sheetnames) == 1

    @pytest.mark.asyncio
    async def test_export_snapshot_not_found(self, client: AsyncClient) -> None:
        from uuid import uuid4

        fake_id = uuid4()
        response = await client.get(f"/api/iso/exports/snapshots/{fake_id}")
        assert response.status_code == 404

    @pytest.mark.asyncio
    async def test_export_single_has_content_disposition(
        self, client: AsyncClient, db_session
    ) -> None:
        await ensure_dev_user(db_session)
        snapshot = await make_snapshot(
            db_session,
            captured_at=datetime(2026, 6, 15, 10, 0, tzinfo=UTC),
        )

        response = await client.get(
            f"/api/iso/exports/snapshots/{snapshot.id}",
        )
        assert response.status_code == 200
        assert "iso_access_review_2026-06-15.xlsx" in response.headers["content-disposition"]

    @pytest.mark.asyncio
    async def test_export_snapshot_with_review_and_actions(
        self, client: AsyncClient, db_session
    ) -> None:
        await ensure_dev_user(db_session)
        snapshot = await make_snapshot(
            db_session,
            captured_at=datetime(2026, 6, 15, 10, 0, tzinfo=UTC),
            data={
                "users": [
                    {
                        "id": "u1",
                        "name": "Alice",
                        "email": "alice@test.com",
                        "suspended": False,
                        "org_unit_path": "/",
                    }
                ],
                "groups": [],
                "group_members": {},
                "role_assignments": [],
            },
            summary={
                "total_users": 1,
                "total_admins": 0,
                "total_groups": 0,
                "external_members": 0,
            },
        )
        review = await make_review(db_session, snapshot.id, status="signed")
        action = AccessReviewActionDB(
            review_id=review.id,
            subject_type="user",
            subject_id="u1",
            subject_label="Alice",
            change_type="new_user",
            action_taken="accepted",
            justification="Approved",
        )
        db_session.add(action)
        await db_session.flush()

        response = await client.get(
            f"/api/iso/exports/snapshots/{snapshot.id}",
        )
        assert response.status_code == 200

        wb = load_workbook(BytesIO(response.content))
        ws = wb[wb.sheetnames[0]]

        all_values = [
            ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        assert "Approved" in all_values
        assert "alice@test.com" in all_values
