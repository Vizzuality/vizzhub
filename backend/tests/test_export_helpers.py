"""Tests for export XLSX helpers."""

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.config import ScoringConfig
from app.core.services.export_helpers import (
    apply_header_style,
    apply_indicator_traffic_light,
    apply_score_traffic_light,
    format_month_header,
    set_column_widths,
)
from app.modules.scorecard.services.export_helpers import create_methodology_sheet
from tests.conftest import load_config_from_csv


@pytest.fixture
def wb():
    return Workbook()


@pytest.fixture
def config():
    return ScoringConfig(load_config_from_csv())


class TestScoreTrafficLight:
    def test_green_above_green_threshold(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=85)
        apply_score_traffic_light(cell, 85, green=80, yellow=60)
        assert cell.fill.start_color.rgb == "00C1E3C2"

    def test_yellow_between_thresholds(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=65)
        apply_score_traffic_light(cell, 65, green=80, yellow=60)
        assert cell.fill.start_color.rgb == "00FFE9A8"

    def test_red_below_yellow_threshold(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=55)
        apply_score_traffic_light(cell, 55, green=80, yellow=60)
        assert cell.fill.start_color.rgb == "00FBBDB9"

    def test_no_fill_when_none(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=None)
        apply_score_traffic_light(cell, None, green=80, yellow=60)
        assert cell.fill == PatternFill()

    def test_exact_boundary_green(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=80)
        apply_score_traffic_light(cell, 80, green=80, yellow=60)
        assert cell.fill.start_color.rgb == "00C1E3C2"

    def test_exact_boundary_yellow(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=60)
        apply_score_traffic_light(cell, 60, green=80, yellow=60)
        assert cell.fill.start_color.rgb == "00FFE9A8"


class TestIndicatorTrafficLight:
    def test_green_above_threshold(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=0.85)
        apply_indicator_traffic_light(cell, 0.85, green=0.8, yellow=0.6)
        assert cell.fill.start_color.rgb == "00E4F3E5"

    def test_yellow_between_thresholds(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=0.65)
        apply_indicator_traffic_light(cell, 0.65, green=0.8, yellow=0.6)
        assert cell.fill.start_color.rgb == "00FFF6DA"

    def test_red_below_threshold(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=0.45)
        apply_indicator_traffic_light(cell, 0.45, green=0.8, yellow=0.6)
        assert cell.fill.start_color.rgb == "00FDE3E1"

    def test_no_fill_when_none(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=None)
        apply_indicator_traffic_light(cell, None)
        assert cell.fill == PatternFill()

    def test_exact_boundary_green(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=0.8)
        apply_indicator_traffic_light(cell, 0.8, green=0.8, yellow=0.6)
        assert cell.fill.start_color.rgb == "00E4F3E5"

    def test_exact_boundary_yellow(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=0.6)
        apply_indicator_traffic_light(cell, 0.6, green=0.8, yellow=0.6)
        assert cell.fill.start_color.rgb == "00FFF6DA"


class TestHeaderStyle:
    def test_applies_bold_and_fill(self, wb):
        ws = wb.active
        ws.append(["Col1", "Col2", "Col3"])
        apply_header_style(ws, row=1)
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            assert cell.font.bold is True


class TestFormatMonthHeader:
    def test_formats_correctly(self):
        assert format_month_header(2025, 1) == "Jan 2025"
        assert format_month_header(2025, 12) == "Dec 2025"


class TestMethodologySheet:
    def test_creates_sheet_with_content(self, wb, config):
        create_methodology_sheet(wb, config)
        assert "Methodology" in wb.sheetnames
        ws = wb["Methodology"]
        assert ws.cell(row=1, column=1).value is not None

    def test_traffic_light_legend_has_categories(self, wb, config):
        ws = create_methodology_sheet(wb, config)
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([v for v in row if v])
        assert "Good performance" in all_values
        assert "At risk / needs attention" in all_values
        assert "Critical / poor performance" in all_values

    def test_traffic_light_legend_has_thresholds(self, wb, config):
        ws = create_methodology_sheet(wb, config)
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(v) for v in row if v])
        assert "Score >= 80" in all_values
        assert "Score >= 60" in all_values
        assert "Score < 60" in all_values

    def test_traffic_light_legend_has_indicator_thresholds(self, wb, config):
        ws = create_methodology_sheet(wb, config)
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(v) for v in row if v])
        assert "Value >= 0.80" in all_values
        assert "Value >= 0.60" in all_values
        assert "Value < 0.60" in all_values


class TestColumnWidths:
    def test_sets_widths(self, wb):
        ws = wb.active
        ws.append(["Name", "Description", "Formula", "Target", "Jan 2025"])
        widths = {"A": 30, "B": 50, "C": 40, "D": 12}
        set_column_widths(ws, widths)
        assert ws.column_dimensions["A"].width == 30
