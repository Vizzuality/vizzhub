from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

FIXTURE_PATH = Path(__file__).resolve().parents[2] / "fixtures" / "accrual_minimal.xlsx"


@pytest.fixture(scope="session", autouse=False)
def _ensure_fixture():
    """Generate the minimal accrual spreadsheet fixture once per session.

    Mirrors the real workbook layout: sheet 'Sales', year row 5, month row 6,
    data rows from row 7. 24 monthly columns (2024 + 2025).
    """
    FIXTURE_PATH.parent.mkdir(parents=True, exist_ok=True)
    if FIXTURE_PATH.exists():
        return FIXTURE_PATH

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Year row 5: cols 13-36
    for i, year in enumerate([2024] * 12 + [2025] * 12):
        ws.cell(row=5, column=13 + i, value=year)

    # Header row 6: cols 1-11 + month numbers cols 13-36
    headers = [
        "Type",
        "Win %",
        "Code",
        "PM",
        "Contract name",
        "Value",
        "Rate",
        "Value €",
        "Start",
        "End",
        "Duration",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=6, column=i, value=h)
    ws.cell(row=6, column=12, value="CHECK")
    for i, month in enumerate(list(range(1, 13)) * 2):
        ws.cell(row=6, column=13 + i, value=month)

    def _write_row(r, type_, code, pm, name, value, rate, value_eur, start, end, duration, monthly):
        ws.cell(row=r, column=1, value=type_)
        ws.cell(row=r, column=3, value=code)
        ws.cell(row=r, column=4, value=pm)
        ws.cell(row=r, column=5, value=name)
        ws.cell(row=r, column=6, value=float(value))
        ws.cell(row=r, column=7, value=float(rate))
        ws.cell(row=r, column=8, value=float(value_eur))
        ws.cell(row=r, column=9, value=start)
        ws.cell(row=r, column=10, value=end)
        ws.cell(row=r, column=11, value=duration)
        # Monthly cells
        for (y, m), amount in monthly.items():
            # year col offset: 2024 = 0..11, 2025 = 12..23
            year_offset = 0 if y == 2024 else 12
            col = 13 + year_offset + (m - 1)
            ws.cell(row=r, column=col, value=float(amount))

    # Contract A: 2024 only, no override.
    monthly_a = {(2024, m): Decimal("1000") for m in range(1, 13)}
    _write_row(
        7,
        "3-Finished",
        "A001",
        "Foo",
        "Contract A",
        12000,
        1.0,
        12000,
        datetime(2024, 1, 1),
        datetime(2024, 12, 1),
        12,
        monthly_a,
    )

    # Contract B: 24 months, with one override (2024-06 = 2000).
    monthly_b = {(y, m): Decimal("1000") for y in (2024, 2025) for m in range(1, 13)}
    monthly_b[(2024, 6)] = Decimal("2000")
    _write_row(
        8,
        "2-Live",
        "B001",
        "Bar",
        "Contract B",
        24000,
        1.1,
        21818.18,
        datetime(2024, 1, 1),
        datetime(2025, 12, 1),
        24,
        monthly_b,
    )

    # Contract C: 2025 H1.
    monthly_c = {(2025, m): Decimal("1000") for m in range(1, 7)}
    _write_row(
        9,
        "3-Finished",
        "C001",
        None,
        "Contract C",
        6000,
        1.05,
        5714.29,
        datetime(2025, 1, 1),
        datetime(2025, 6, 1),
        6,
        monthly_c,
    )

    wb.save(FIXTURE_PATH)
    return FIXTURE_PATH
