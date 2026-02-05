"""Tests for ExportService."""

import pytest
import pytest_asyncio
from datetime import date
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import ScoringConfig
from app.models.global_metrics import GlobalMetricsDB
from app.models.metrics import MetricsDB
from app.models.project import ProjectDB
from app.services.export_service import ExportService


@pytest_asyncio.fixture
async def project_with_3_months(
    db_session: AsyncSession, scoring_config: ScoringConfig
) -> ProjectDB:
    """Create a project with 3 months of cumulative metrics."""
    project = ProjectDB(
        id=uuid4(),
        name="Export Test Project",
        jira_project_key="EXP",
        github_repo="test/export",
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
        status="in_progress",
    )
    db_session.add(project)
    await db_session.flush()

    for month in range(1, 4):
        metrics = MetricsDB(
            project_id=project.id,
            period_start=date(2025, month, 1),
            period_end=date(2025, month, 28),
            period_year=2025,
            period_month=month,
            snapshot_type="cumulative",
            weights_applied={},
            targets_applied={},
            budget_total=Decimal("100000"),
            cost_to_date=Decimal(str(40000 + month * 5000)),
            percent_completed=Decimal(str(0.1 * month)),
            percent_planned=Decimal(str(0.1 * month)),
            bugs_total=5 + month,
            tasks_completed=100,
            governance_exceptions=1,
            sev1_incident=False,
            total_merged_prs=50,
            prs_without_review=2,
        )
        db_session.add(metrics)

    await db_session.commit()
    await db_session.refresh(project)
    return project


class TestExportServiceProjectDetail:
    @pytest.mark.asyncio
    async def test_generates_valid_xlsx(
        self, db_session, scoring_config, project_with_3_months
    ):
        project = project_with_3_months
        service = ExportService(scoring_config)
        output = await service.export_project_detail(
            db=db_session,
            project=project,
            start_year=2025,
            start_month=1,
            end_year=2025,
            end_month=3,
            snapshot_type="cumulative",
        )
        assert isinstance(output, BytesIO)
        wb = load_workbook(output)
        assert "Summary" in wb.sheetnames
        assert "Metrics" in wb.sheetnames
        assert "Methodology" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_summary_sheet_has_project_info(
        self, db_session, scoring_config, project_with_3_months
    ):
        project = project_with_3_months
        service = ExportService(scoring_config)
        output = await service.export_project_detail(
            db=db_session,
            project=project,
            start_year=2025,
            start_month=1,
            end_year=2025,
            end_month=3,
            snapshot_type="cumulative",
        )
        wb = load_workbook(output)
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 10)]
        assert "Export Test Project" in values

    @pytest.mark.asyncio
    async def test_metrics_sheet_has_month_columns(
        self, db_session, scoring_config, project_with_3_months
    ):
        project = project_with_3_months
        service = ExportService(scoring_config)
        output = await service.export_project_detail(
            db=db_session,
            project=project,
            start_year=2025,
            start_month=1,
            end_year=2025,
            end_month=3,
            snapshot_type="cumulative",
        )
        wb = load_workbook(output)
        ws = wb["Metrics"]
        headers = [ws.cell(row=1, column=c).value for c in range(5, 8)]
        assert "Jan 2025" in headers
        assert "Feb 2025" in headers
        assert "Mar 2025" in headers

    @pytest.mark.asyncio
    async def test_metrics_sheet_has_hierarchical_rows(
        self, db_session, scoring_config, project_with_3_months
    ):
        project = project_with_3_months
        service = ExportService(scoring_config)
        output = await service.export_project_detail(
            db=db_session,
            project=project,
            start_year=2025,
            start_month=1,
            end_year=2025,
            end_month=3,
            snapshot_type="cumulative",
        )
        wb = load_workbook(output)
        ws = wb["Metrics"]
        assert ws.cell(row=2, column=1).value == "FINAL SCORE"

    @pytest.mark.asyncio
    async def test_empty_range_returns_xlsx_with_no_data_columns(
        self, db_session, scoring_config, project_with_3_months
    ):
        project = project_with_3_months
        service = ExportService(scoring_config)
        output = await service.export_project_detail(
            db=db_session,
            project=project,
            start_year=2024,
            start_month=1,
            end_year=2024,
            end_month=3,
            snapshot_type="cumulative",
        )
        assert isinstance(output, BytesIO)
        wb = load_workbook(output)
        assert "Metrics" in wb.sheetnames


@pytest_asyncio.fixture
async def global_metrics_3_months(db_session: AsyncSession) -> list[GlobalMetricsDB]:
    """Create 3 months of pre-computed global metrics."""
    records = []
    for month in range(1, 4):
        record = GlobalMetricsDB(
            period_year=2025,
            period_month=month,
            project_count=3,
            score=75.0 + month,
            score_count=3,
            p_time=80.0,
            p_time_count=3,
            p_cost=70.0,
            p_cost_count=2,
            p_quality=65.0 + month,
            p_quality_count=3,
            p_value=60.0,
            p_value_count=1,
            p_satisfaction=85.0,
            p_satisfaction_count=3,
            p_flow=72.0,
            p_flow_count=3,
            p_engineering=78.0,
            p_engineering_count=3,
            p_risk=90.0,
            p_risk_count=3,
            spi=0.95,
            spi_count=3,
            cpi=0.88,
            cpi_count=2,
            on_time_milestones=0.8,
            on_time_milestones_count=3,
            defect_density=0.03,
            defect_density_count=3,
            pr_review_ratio=0.96,
            pr_review_ratio_count=3,
            governance_compliance=1.0,
            governance_compliance_count=3,
        )
        db_session.add(record)
        records.append(record)
    await db_session.commit()
    return records


class TestExportServiceGlobalDashboard:
    @pytest.mark.asyncio
    async def test_generates_valid_xlsx(
        self, db_session, scoring_config, global_metrics_3_months
    ):
        service = ExportService(scoring_config)
        output = await service.export_global_dashboard(
            db=db_session,
            start_year=2025,
            start_month=1,
            end_year=2025,
            end_month=3,
            snapshot_type="cumulative",
        )
        assert isinstance(output, BytesIO)
        wb = load_workbook(output)
        assert "Summary" in wb.sheetnames
        assert "Metrics" in wb.sheetnames
        assert "Methodology" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_summary_has_scores_and_dimensions(
        self, db_session, scoring_config, global_metrics_3_months
    ):
        service = ExportService(scoring_config)
        output = await service.export_global_dashboard(
            db=db_session,
            start_year=2025,
            start_month=1,
            end_year=2025,
            end_month=3,
            snapshot_type="cumulative",
        )
        wb = load_workbook(output)
        ws = wb["Summary"]
        row_labels = [ws.cell(row=r, column=1).value for r in range(1, 15)]
        assert "Projects" in row_labels
        assert "Overall Score" in row_labels

    @pytest.mark.asyncio
    async def test_summary_contains_project_count(
        self, db_session, scoring_config, global_metrics_3_months
    ):
        service = ExportService(scoring_config)
        output = await service.export_global_dashboard(
            db=db_session,
            start_year=2025,
            start_month=1,
            end_year=2025,
            end_month=1,
            snapshot_type="cumulative",
        )
        wb = load_workbook(output)
        ws = wb["Summary"]
        # Row 4 = Projects, Col 2 = first month value
        assert ws.cell(row=4, column=2).value == 3

    @pytest.mark.asyncio
    async def test_metrics_sheet_has_hierarchical_rows(
        self, db_session, scoring_config, global_metrics_3_months
    ):
        service = ExportService(scoring_config)
        output = await service.export_global_dashboard(
            db=db_session,
            start_year=2025,
            start_month=1,
            end_year=2025,
            end_month=3,
            snapshot_type="cumulative",
        )
        wb = load_workbook(output)
        ws = wb["Metrics"]
        assert ws.cell(row=2, column=1).value == "FINAL SCORE"

    @pytest.mark.asyncio
    async def test_metrics_sheet_has_indicator_values(
        self, db_session, scoring_config, global_metrics_3_months
    ):
        service = ExportService(scoring_config)
        output = await service.export_global_dashboard(
            db=db_session,
            start_year=2025,
            start_month=1,
            end_year=2025,
            end_month=1,
            snapshot_type="cumulative",
        )
        wb = load_workbook(output)
        ws = wb["Metrics"]
        # Collect all values in column 5 (first month data column)
        values = [ws.cell(row=r, column=5).value for r in range(2, ws.max_row + 1)]
        non_none = [v for v in values if v is not None]
        assert len(non_none) > 0

    @pytest.mark.asyncio
    async def test_no_data_returns_valid_xlsx(
        self, db_session, scoring_config
    ):
        service = ExportService(scoring_config)
        output = await service.export_global_dashboard(
            db=db_session,
            start_year=2025,
            start_month=1,
            end_year=2025,
            end_month=3,
            snapshot_type="cumulative",
        )
        assert isinstance(output, BytesIO)
        wb = load_workbook(output)
        assert "Summary" in wb.sheetnames
        assert "Metrics" in wb.sheetnames


class TestExportServiceHelpers:
    def test_generate_periods_single_month(self, scoring_config):
        service = ExportService(scoring_config)
        periods = service._generate_periods(2025, 3, 2025, 3)
        assert periods == [(2025, 3)]

    def test_generate_periods_cross_year(self, scoring_config):
        service = ExportService(scoring_config)
        periods = service._generate_periods(2024, 11, 2025, 2)
        assert periods == [(2024, 11), (2024, 12), (2025, 1), (2025, 2)]

    def test_extract_value_none_data(self, scoring_config):
        service = ExportService(scoring_config)
        assert service._extract_value("final_score", 0, None) is None

    def test_parse_target_dash(self, scoring_config):
        assert ExportService._parse_target("-") is None

    def test_parse_target_valid(self, scoring_config):
        assert ExportService._parse_target("80") == 80.0

    def test_parse_target_none(self, scoring_config):
        assert ExportService._parse_target(None) is None
