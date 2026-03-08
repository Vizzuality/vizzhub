"""Tests for ISO export service — GitHub provider XLSX generation."""

from datetime import datetime, timezone
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook

from app.modules.iso.services.export_service import IsoExportService


def _make_github_snapshot(
    captured_at=None,
    org="acme-corp",
    members=None,
    teams=None,
    team_members=None,
    outside_collaborators=None,
):
    """Build a fake GitHub snapshot dict matching AccessSnapshotDB shape."""
    members = members or [
        {"login": "alice", "name": "Alice A", "email": "alice@co.com", "role": "admin"},
        {"login": "bob", "name": "Bob B", "email": None, "role": "member"},
    ]
    teams = teams or [
        {
            "name": "Backend",
            "slug": "backend",
            "parent_slug": None,
            "privacy": "closed",
        },
    ]
    team_members = team_members or {
        "backend": [
            {"login": "alice", "role": "maintainer"},
            {"login": "bob", "role": "member"},
        ],
    }
    outside_collaborators = outside_collaborators or [
        {"login": "contractor1", "name": "Contractor", "email": "c@ext.com"},
    ]

    return {
        "id": str(uuid4()),
        "provider": "github",
        "captured_at": captured_at or datetime(2026, 6, 15, 10, 0, tzinfo=timezone.utc),
        "data_version": "2",
        "source_metadata": {"org": org, "collector": "github"},
        "data": {
            "members": members,
            "teams": teams,
            "team_members": team_members,
            "outside_collaborators": outside_collaborators,
        },
        "summary": {
            "total_members": len(members),
            "total_admins": sum(1 for m in members if m["role"] == "admin"),
            "total_teams": len(teams),
            "outside_collaborators": len(outside_collaborators),
        },
    }


def _make_gw_snapshot(captured_at=None, domain="test.com"):
    """Build a minimal Google Workspace snapshot for mixed-provider tests."""
    return {
        "id": str(uuid4()),
        "provider": "google_workspace",
        "captured_at": captured_at or datetime(2026, 6, 15, 10, 0, tzinfo=timezone.utc),
        "data_version": "1",
        "source_metadata": {"domain": domain},
        "data": {
            "users": [
                {
                    "id": "u1",
                    "name": "Alice",
                    "email": "alice@test.com",
                    "suspended": False,
                    "org_unit_path": "/",
                },
            ],
            "groups": [],
            "group_members": {},
            "role_assignments": [],
        },
        "summary": {
            "total_users": 1,
            "total_admins": 0,
            "total_groups": 0,
            "external_members": 0,
        },
    }


def _make_review(status="signed", notes="All good"):
    return {
        "id": str(uuid4()),
        "status": status,
        "scope": "All members",
        "notes": notes,
        "reviewer_email": "admin@acme.com",
        "signed_by_email": "admin@acme.com",
        "signed_at": datetime(2026, 6, 16, 12, 0, tzinfo=timezone.utc),
        "diff_summary": None,
    }


def _all_cell_values(ws) -> list:
    """Flatten all cell values in a worksheet into a single list."""
    return [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    ]


class TestIsoGitHubExport:
    def test_github_snapshot_export(self):
        service = IsoExportService()
        snapshot = _make_github_snapshot()
        review = _make_review()

        output = service.export_snapshots(
            snapshots_with_reviews=[(snapshot, review, [])]
        )

        assert isinstance(output, BytesIO)
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 1

        values = _all_cell_values(wb[wb.sheetnames[0]])
        assert "Members" in values
        assert "Teams" in values
        assert "Outside Collaborators" in values

    def test_github_export_header_shows_org(self):
        service = IsoExportService()
        snapshot = _make_github_snapshot(org="my-github-org")
        review = _make_review()

        output = service.export_snapshots(
            snapshots_with_reviews=[(snapshot, review, [])]
        )
        wb = load_workbook(output)
        ws = wb[wb.sheetnames[0]]

        labels = [ws.cell(row=r, column=1).value for r in range(1, 20)]
        assert "Organization" in labels

        org_row = labels.index("Organization") + 1
        assert ws.cell(row=org_row, column=2).value == "my-github-org"

        assert "Provider" in labels
        provider_row = labels.index("Provider") + 1
        assert ws.cell(row=provider_row, column=2).value == "github"

    def test_github_export_members_with_name_email(self):
        service = IsoExportService()
        snapshot = _make_github_snapshot(
            members=[
                {"login": "alice", "name": "Alice A", "email": "alice@co.com", "role": "admin"},
                {"login": "bob", "name": "Bob B", "email": None, "role": "member"},
            ],
        )

        output = service.export_snapshots(
            snapshots_with_reviews=[(snapshot, None, [])]
        )
        wb = load_workbook(output)
        ws = wb[wb.sheetnames[0]]

        values = _all_cell_values(ws)
        assert "alice" in values
        assert "Alice A" in values
        assert "alice@co.com" in values
        assert "bob" in values
        assert "Bob B" in values
        assert "admin" in values
        assert "member" in values

    def test_github_export_outside_collaborators_with_name_email(self):
        service = IsoExportService()
        snapshot = _make_github_snapshot(
            outside_collaborators=[
                {"login": "ext1", "name": "Ext One", "email": "ext@co.com"},
            ],
        )

        output = service.export_snapshots(
            snapshots_with_reviews=[(snapshot, None, [])]
        )
        wb = load_workbook(output)
        ws = wb[wb.sheetnames[0]]

        values = _all_cell_values(ws)
        assert "ext1" in values
        assert "Ext One" in values
        assert "ext@co.com" in values

    def test_mixed_gw_and_github_export(self):
        service = IsoExportService()
        gw_snap = _make_gw_snapshot(
            captured_at=datetime(2026, 3, 1, tzinfo=timezone.utc),
            domain="acme.com",
        )
        gh_snap = _make_github_snapshot(
            captured_at=datetime(2026, 3, 2, tzinfo=timezone.utc),
            org="acme-corp",
        )
        gw_review = _make_review()
        gh_review = _make_review()

        output = service.export_snapshots(
            snapshots_with_reviews=[
                (gw_snap, gw_review, []),
                (gh_snap, gh_review, []),
            ]
        )
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 2

        gw_values = _all_cell_values(wb[wb.sheetnames[0]])
        assert "acme.com" in gw_values
        assert "google_workspace" in gw_values
        assert "Users" in gw_values

        gh_values = _all_cell_values(wb[wb.sheetnames[1]])
        assert "acme-corp" in gh_values
        assert "github" in gh_values
        assert "Members" in gh_values
