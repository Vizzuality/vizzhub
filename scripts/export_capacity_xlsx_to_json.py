"""Export Capacity management.xlsx 'General view' tab to capacity_seed.json.

Run locally where the xlsx file exists:
    python scripts/export_capacity_xlsx_to_json.py temp/Capacity\ management.xlsx

Outputs: capacity_seed.json in the current directory.
"""

import json
import sys
from datetime import date, timedelta

import openpyxl

# Manual mapping: spreadsheet name → user email.
# Fill in before running. Names not in this dict are skipped with a warning.
NAME_TO_EMAIL: dict[str, str] = {
    "Adam Trincas": "adam.trincas@vizzuality.com",
    "Adélaïde Cadioux": "adelaide.cadioux@vizzuality.com",
    "Alex Larrañaga": "alex.larranaga@vizzuality.com",
    "Alicia Arenzana": "alicia.arenzana@vizzuality.com",
    "Andrea Rota": "andrea.rota@vizzuality.com",
    "Andreia Ribeiro": "andreia.ribeiro@vizzuality.com",
    "Andrés Gonzalez": "andres.gonzalez@vizzuality.com",
    "Angel Arcones": "angel.arcones@vizzuality.com",
    "Angela Minto": "angela.minto@vizzuality.com",
    "Ariadna Ariza": "ariadna.ariza@vizzuality.com",
    "Clara Linos": "clara.linos@vizzuality.com",
    "Clement Prodhomme": "clement.prodhomme@vizzuality.com",
    "Cristina Jaramago": "cristina.jaramago@vizzuality.com",
    "Daniel Caso": "daniel.caso@vizzuality.com",
    "Edgar Espinoza": "edgar.espinoza@vizzuality.com",
    "Elena Palao": "elena.palao@vizzuality.com",
    "Iker Sanchez": "iker.sanchez@vizzuality.com",
    "Irene Rodriguez": "irene.rodriguez@vizzuality.com",
    "Jacinta Hamley": "jacinta.hamley@vizzuality.com",
    "Javi Lois": "javier.lois@vizzuality.com",
    "Kevin Sánchez": "kevin.sanchez@vizzuality.com",
    "Laura Riera": "laura.riera@vizzuality.com",
    "Lyubov Pencheva": "lyubov.pencheva@vizzuality.com",
    "María Luena": "maria.luena@vizzuality.com",
    "María Relea": "maria.relea@vizzuality.com",
    "Miguel Barrenechea": "miguel.barrenechea@vizzuality.com",
    "Biel Stela": "biel.stela@vizzuality.com",
    "Cecilia Fili": "cecilia.fili@vizzuality.com",
    "Miguel Mendoza": "miguel.mendoza@vizzuality.com",
    "Mike Harfoot": "mike.harfoot@vizzuality.com",
    "Santiago Ferrer": "santiago.ferrer@vizzuality.com",
    "Sergio Estella": "sergio.estella@vizzuality.com",
    "Sofia Aldabet": "sofia.aldabet@vizzuality.com",
    "Susana Romao": "susana.romao@vizzuality.com",
}

# Spreadsheet project names → DB project names (where they differ)
PROJECT_NAME_MAPPING: dict[str, str] = {
    "Amazonia 360": "Amazonia360 - MVP",
    "CCSA": "Climate Smart Map",
    "Catalyse": "CATALYSE",
    "Climate Watch": "Climate Watch 2026 Maintenance Contract",
    "ECCC": "ECCC HJBL",
    "ESA-GDA": "ESA GDA Comms 2023 - 2027",
    "FHWPC": "FHWPC Implementation phase",
    "Global Rangelands": "Global Rangelands Data Platform (GRASS) - MVP",
    "ICIMOD": "ICIMOD web discovery",
    "IDB": "IDB ESGeoHub",
    "IIASA": "IIASA Scenario Compass Platform ",
    "Landscape Capital Explorer": "Landscape Capital Explorer - Implementation Phase",
    "MIRACA": "HE MIRACA",
    "Mars": "Mars MGIS Maintenance 2025",
    "Marxan": "Marxan maintenance 2025-2028",
    "Open Earth Monitor (OEM)": "Open Earth Monitor ",
    "South Sudán": "South Sudan Visualization Pilot",
    "Unesco": "UNESCO WHOMP",
    "WB Country Workspace": "World bank country workspace MVP design",
    "Wetlands - Gap Map": "Wetlands gap map",
}

DATA_START_ROW = 16
PROJECT_COL = 1
ROLE_COL = 2
NAME_COL = 3
WEEK_START_COL = 4

# Columns 4-11 use "days" format (1-5), columns 12+ use percentage (0-100+)
DAYS_FORMAT_END_COL = 11


def iso_week_to_monday(year: int, week_num: int) -> date:
    """Convert ISO year + week number to the Monday of that week."""
    jan1 = date(year, 1, 1)
    # Find the Monday of ISO week 1
    day_of_week = jan1.isoweekday()
    if day_of_week <= 4:
        iso_week1_monday = jan1 - timedelta(days=day_of_week - 1)
    else:
        iso_week1_monday = jan1 + timedelta(days=8 - day_of_week)
    return iso_week1_monday + timedelta(weeks=week_num - 1)


def _build_month_to_year(ws, max_col: int) -> dict[str, int]:
    """Read month headers (row 14) to determine year boundaries."""
    month_to_year: dict[str, int] = {}
    for col in range(WEEK_START_COL, max_col + 1):
        month_name = ws.cell(row=14, column=col).value
        if not month_name:
            continue
        if month_name in ("November", "December"):
            month_to_year[month_name] = 2025
        else:
            month_to_year[month_name] = 2026
    return month_to_year


def _build_week_info(ws, max_col: int, month_to_year: dict[str, int]) -> dict[int, dict]:
    """Build week_num -> Monday mapping from header rows."""
    week_info: dict[int, dict] = {}
    current_year = 2025
    for col in range(WEEK_START_COL, max_col + 1):
        week_num_val = ws.cell(row=12, column=col).value
        month_val = ws.cell(row=14, column=col).value
        if month_val and month_val in month_to_year:
            current_year = month_to_year[month_val]
        if week_num_val is None:
            continue
        week_num = int(week_num_val)
        if week_num < 10 and current_year == 2025:
            current_year = 2026
        monday = iso_week_to_monday(current_year, week_num)
        week_info[col] = {
            "week_num": week_num,
            "monday": monday,
            "is_days_format": col <= DAYS_FORMAT_END_COL,
        }
    return week_info


def _parse_cell_percentage(val, is_days_format: bool) -> int | None:
    """Parse a cell value to percentage. Returns None if invalid/empty."""
    if val is None or val == "" or str(val).lower() == "x":
        return None
    try:
        num = float(val)
    except (ValueError, TypeError):
        return None
    if num <= 0:
        return None
    percentage = int(num * 20) if is_days_format else int(num)
    if percentage < 1:
        return None
    return min(percentage, 200)


def _extract_data_rows(ws, week_info: dict[int, dict]) -> tuple[list[dict], set[str]]:
    """Read data rows and return (records, skipped_names)."""
    records: list[dict] = []
    skipped_names: set[str] = set()

    for row in range(DATA_START_ROW, ws.max_row + 1):
        person_name = ws.cell(row=row, column=NAME_COL).value
        if not person_name:
            continue

        person_name = str(person_name).strip()
        email = NAME_TO_EMAIL.get(person_name)
        if not email:
            skipped_names.add(person_name)
            continue

        raw_project = ws.cell(row=row, column=PROJECT_COL).value
        project_name = str(raw_project).strip() if raw_project else None
        if not project_name:
            continue
        project_name = PROJECT_NAME_MAPPING.get(project_name, project_name)

        for col, info in week_info.items():
            percentage = _parse_cell_percentage(
                ws.cell(row=row, column=col).value, info["is_days_format"],
            )
            if percentage is None:
                continue
            records.append({
                "project_name": project_name,
                "user_email": email,
                "week_start": info["monday"].isoformat(),
                "percentage": percentage,
            })

    return records, skipped_names


def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["General view"]

    month_to_year = _build_month_to_year(ws, ws.max_column)
    week_info = _build_week_info(ws, ws.max_column, month_to_year)
    records, skipped_names = _extract_data_rows(ws, week_info)

    output_path = "capacity_seed.json"
    with open(output_path, "w") as f:
        json.dump(records, f, indent=2)

    print(f"Exported {len(records)} records to {output_path}")
    if skipped_names:
        print(f"Skipped names (no email mapping): {sorted(skipped_names)}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f"Usage: python {sys.argv[0]} <path-to-xlsx>")
        sys.exit(1)
    main(sys.argv[1])
